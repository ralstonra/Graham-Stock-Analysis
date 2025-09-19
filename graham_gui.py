import tkinter as tk
from tkinter import ttk, scrolledtext, messagebox, filedialog, simpledialog
import asyncio
import logging
import pandas as pd
import sqlite3
import sys
import threading
import time
import os
import json
import yfinance as yf
from decouple import config
from datetime import datetime, timedelta
from graham_data import (screen_exchange_graham_stocks, fetch_batch_data,
                         fetch_stock_data, get_stocks_connection, fetch_with_multiple_keys_async,
                         NYSE_LIST_FILE, NASDAQ_LIST_FILE, TickerManager, get_file_hash,
                         calculate_graham_value, calculate_graham_score_8, calculate_common_criteria, 
                         clear_in_memory_caches, save_qualifying_stocks_to_favorites, get_stock_data_from_db,
                         get_sector_growth_rate, calculate_cagr, get_aaa_yield, get_bank_metrics,
                         get_tangible_book_value_per_share)
from config import (
    BASE_DIR, FMP_API_KEYS, FAVORITES_FILE, paid_rate_limiter, free_rate_limiter, 
    CACHE_EXPIRY, screening_logger, analyze_logger, USER_DATA_DIR, FRED_API_KEY
)
import queue
import shutil
import requests
import ftplib
import openpyxl
import re
import subprocess
import platform
from openpyxl.chart import LineChart, Reference, Series, ScatterChart
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.chart.marker import Marker
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.chart.axis import ChartLines
from openpyxl.formatting.rule import FormulaRule, CellIsRule

FAVORITES_LOCK = threading.Lock()
DATA_DIR = BASE_DIR

class GrahamScreeningApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Stock Analysis (Graham Defensive)")
        self.root.geometry("1200x900")

        # Variables
        self.nyse_screen_var = tk.BooleanVar(value=False)
        self.nasdaq_screen_var = tk.BooleanVar(value=False)
        self.financial_screen_var = tk.BooleanVar(value=False)
        self.cancel_event = threading.Event()
        self.margin_of_safety_var = tk.DoubleVar(value=33.0)
        self.expected_return_var = tk.DoubleVar(value=0.0)
        self.ticker_manager = TickerManager(NYSE_LIST_FILE, NASDAQ_LIST_FILE, USER_DATA_DIR)
        self.screening_active = False
        self.analysis_lock = threading.Lock()
        self.ticker_cache = {}
        self.ticker_cache_lock = threading.Lock()

        # Asyncio Thread Setup
        self.task_queue = queue.Queue()
        self.asyncio_thread = threading.Thread(target=self.run_asyncio_loop, args=(self.task_queue,), daemon=True)
        self.asyncio_thread.start()

        # Main Frame Layout
        self.main_frame = ttk.Frame(root)
        self.main_frame.pack(expand=True, fill="both", padx=5, pady=5)
        self.main_frame.grid_rowconfigure(0, weight=3)
        self.main_frame.grid_rowconfigure(1, weight=1)
        self.main_frame.grid_columnconfigure((0, 1, 2), weight=1, minsize=400)

        # Sub-Frames
        self.left_frame = ttk.Frame(self.main_frame, width=400)
        self.middle_frame = ttk.Frame(self.main_frame, width=400)
        self.right_frame = ttk.Frame(self.main_frame, width=400)

        self.left_frame.grid(row=0, column=0, sticky="nsew", padx=2)
        self.middle_frame.grid(row=0, column=1, sticky="nsew", padx=2)
        self.right_frame.grid(row=0, column=2, sticky="nsew", padx=2)

        self.left_frame.grid_columnconfigure(0, weight=1)
        self.left_frame.grid_rowconfigure(tuple(range(19)), weight=1)
        self.middle_frame.grid_columnconfigure(0, weight=1)
        self.right_frame.grid_columnconfigure(0, weight=1)

        # Styling
        style = ttk.Style()
        style.theme_use("clam")
        style.configure("TLabel", font=("Helvetica", 9), background="#f0f0f0")
        style.configure("TButton", font=("Helvetica", 9), padding=2)
        style.configure("TCheckbutton", font=("Helvetica", 9))
        style.configure("TEntry", font=("Helvetica", 9))
        style.configure("TCombobox", font=("Helvetica", 9))
        style.configure("TProgressbar", thickness=15)

        self.create_widgets()

        paid_rate_limiter.on_sleep = self.on_rate_limit_sleep

        # New: Tooltips
        self.create_tooltip(self.entry, "Enter comma-separated tickers, e.g., AOS, AAPL")
        self.create_tooltip(self.search_entry, "Search by ticker or company name")
        self.create_tooltip(self.favorite_menu, "Select a saved favorites list")
        self.create_tooltip(self.margin_of_safety_label, "Discount to intrinsic value for safety margin")
        self.create_tooltip(self.expected_return_label, "Premium above intrinsic value for sell target")
        self.create_tooltip(self.analyze_button, "Analyze the entered tickers")
        self.create_tooltip(self.progress_bar, "Shows progress of screening or analysis")

        self.check_for_updates()

    # New: Helper method for tooltips
    def create_tooltip(self, widget: tk.Widget, text: str) -> None:
        tooltip = tk.Toplevel(self.root)
        tooltip.withdraw()
        tooltip.wm_overrideredirect(True)
        label = tk.Label(tooltip, text=text, background="yellow", relief="solid", borderwidth=1, padx=5, pady=3)
        label.pack()

        def show_tooltip(event):
            x, y, _, _ = widget.bbox("insert")
            x += widget.winfo_rootx() + 25
            y += widget.winfo_rooty() + 25
            tooltip.wm_geometry(f"+{x}+{y}")
            tooltip.deiconify()

        def hide_tooltip(event):
            tooltip.withdraw()

        widget.bind("<Enter>", show_tooltip)
        widget.bind("<Leave>", hide_tooltip)

    def on_rate_limit_sleep(self, sleep_time):
        message = f"Rate limit reached, pausing for {sleep_time / 60:.1f} minutes"
        self.root.after(0, lambda: self.rate_limit_label.config(text=message))

    def run_asyncio_loop(self, task_queue):
        loop = asyncio.new_event_loop()
        asyncio.set_event_loop(loop)
        while True:
            coro = task_queue.get()
            if coro is None:
                break
            loop.run_until_complete(coro)
        loop.close()

    def create_widgets(self):
        # Left Frame Widgets (Analyze Stocks)
        self.left_frame.grid_rowconfigure(tuple(range(15)), weight=1)

        ttk.Label(self.left_frame, text="Enter Stock Tickers (comma-separated, e.g., AOS, AAPL):").grid(row=0, column=0, pady=2, sticky="ew")
        self.entry = ttk.Entry(self.left_frame, width=50)
        self.entry.grid(row=1, column=0, pady=2, sticky="ew")
        self.entry.bind('<FocusOut>', lambda e: self.validate_tickers())

        ttk.Label(self.left_frame, text="Search Results:").grid(row=2, column=0, pady=2, sticky="ew")
        self.search_entry = ttk.Entry(self.left_frame, width=50)
        self.search_entry.grid(row=3, column=0, pady=2, sticky="ew")
        self.search_entry.bind('<KeyRelease>', self.filter_tree)

        self.favorites = self.load_favorites()
        ttk.Label(self.left_frame, text="Favorite Lists:").grid(row=4, column=0, pady=2, sticky="ew")
        self.favorite_var = tk.StringVar(value="Select Favorite")
        self.favorite_menu = ttk.Combobox(self.left_frame, textvariable=self.favorite_var, values=list(self.favorites.keys()))
        self.favorite_menu.grid(row=5, column=0, pady=2, sticky="ew")

        def load_favorite(event=None):
            selected = self.favorite_var.get()
            if selected and selected != "Select Favorite":
                self.entry.delete(0, tk.END)
                tickers = self.favorites[selected]
                self.entry.insert(0, ",".join(tickers))

        self.favorite_menu.bind('<<ComboboxSelected>>', load_favorite)

        def save_favorite():
            print("Save Favorite button clicked")  # Debug
            self.task_queue.put(self.ticker_manager.initialize())
            time.sleep(1)  # Wait for initialization (adjust if needed)
            print(f"NYSE tickers: {len(self.ticker_manager.nyse_tickers)}, NASDAQ tickers: {len(self.ticker_manager.nasdaq_tickers)}")  # Debug
            tickers_input = self.entry.get()
            print(f"Input: {tickers_input}")  # Debug
            if not self.validate_tickers():
                messagebox.showwarning("Invalid Tickers", "No valid tickers entered.")
                return
            name = simpledialog.askstring("Save Favorite", "Enter list name:")
            print(f"List name: {name}")  # Debug
            if not name or not name.strip():
                messagebox.showwarning("Invalid Name", "Please enter a valid list name.")
                return
            if name and tickers_input.strip():
                tickers = self.parse_tickers(tickers_input)
                print(f"Parsed tickers: {tickers}")  # Debug
                valid_tickers = [t for t in tickers if self.ticker_manager.is_valid_ticker(t)]
                print(f"Valid tickers: {valid_tickers}")  # Debug
                if not valid_tickers:
                    messagebox.showwarning("Invalid Tickers", "No valid NYSE or NASDAQ tickers.")
                    return
                self.favorites[name] = valid_tickers
                self.save_favorites()
                self.refresh_favorites_dropdown(name)
                print(f"Saved favorites: {self.favorites}")  # Debug

        ttk.Button(self.left_frame, text="Save Favorite", command=save_favorite).grid(row=6, column=0, pady=2, sticky="ew")
        ttk.Button(self.left_frame, text="Manage Favorites", command=self.manage_favorites).grid(row=7, column=0, pady=2, sticky="ew")

        # Margin of Safety
        self.margin_of_safety_label = ttk.Label(self.left_frame, text="Margin of Safety: 33%")
        self.margin_of_safety_label.grid(row=8, column=0, pady=2, sticky="ew")
        ttk.Scale(self.left_frame, from_=0, to=50, orient=tk.HORIZONTAL, variable=self.margin_of_safety_var,
                  command=lambda value: self.margin_of_safety_label.config(text=f"Margin of Safety: {int(float(value))}%")).grid(row=9, column=0, pady=2, sticky="ew")

        # Expected Rate of Return
        self.expected_return_label = ttk.Label(self.left_frame, text="Expected Rate of Return: 0%")
        self.expected_return_label.grid(row=10, column=0, pady=2, sticky="ew")
        ttk.Scale(self.left_frame, from_=0, to=20, orient=tk.HORIZONTAL, variable=self.expected_return_var,
                  command=lambda value: self.expected_return_label.config(text=f"Expected Rate of Return: {int(float(value))}%")).grid(row=11, column=0, pady=2, sticky="ew")

        self.analyze_button = ttk.Button(self.left_frame, text="Analyze Stocks", command=self.analyze_multiple_stocks)
        self.analyze_button.grid(row=14, column=0, pady=2, sticky="ew")

        # Middle Frame Widgets (Utility Buttons and Progress/Status)
        ttk.Button(self.middle_frame, text="Check Ticker File Ages", command=self.check_file_age).grid(row=0, column=0, pady=2, sticky="ew")
        ttk.Button(self.middle_frame, text="Update Ticker Files", command=self.update_ticker_files).grid(row=1, column=0, pady=2, sticky="ew")
        ttk.Button(self.middle_frame, text="Clear Cache", command=self.clear_cache).grid(row=2, column=0, pady=2, sticky="ew")
        ttk.Button(self.middle_frame, text="Help", command=self.show_help).grid(row=3, column=0, pady=2, sticky="ew")
        ttk.Button(self.middle_frame, text="Cancel Screening", command=self.cancel_screening).grid(row=4, column=0, pady=2, sticky="ew")

        self.progress_var = tk.DoubleVar()
        self.progress_bar = ttk.Progressbar(self.middle_frame, variable=self.progress_var, maximum=100, length=300, mode='determinate')
        self.progress_bar.grid(row=5, column=0, pady=2, sticky="ew")
        self.progress_label = ttk.Label(self.middle_frame, text="Progress: 0% (0/0 stocks processed, 0 passed)")
        self.progress_label.grid(row=6, column=0, pady=2, sticky="ew")

        self.cache_var = tk.DoubleVar()
        self.cache_bar = ttk.Progressbar(self.middle_frame, variable=self.cache_var, maximum=100, length=300, mode='determinate')
        self.cache_bar.grid(row=7, column=0, pady=2, sticky="ew")
        self.cache_label = ttk.Label(self.middle_frame, text="Cache Usage: 0% (0 cached, 0 fresh)")
        self.cache_label.grid(row=8, column=0, pady=2, sticky="ew")

        self.rate_limit_label = ttk.Label(self.middle_frame, text="")
        self.rate_limit_label.grid(row=9, column=0, pady=2, sticky="ew")
        self.status_label = ttk.Label(self.middle_frame, text="")
        self.status_label.grid(row=10, column=0, pady=2, sticky="ew")

        # Right Frame Widgets (Screenings)
        ttk.Checkbutton(self.right_frame, text="Run NYSE Graham Screening", variable=self.nyse_screen_var).grid(row=0, column=0, pady=2, padx=2, sticky="w")
        ttk.Button(self.right_frame, text="Run NYSE Screening", command=self.run_nyse_screening).grid(row=1, column=0, pady=2, sticky="ew")
        ttk.Button(self.right_frame, text="Show NYSE Qualifying Stocks", command=self.display_nyse_qualifying_stocks).grid(row=2, column=0, pady=2, sticky="ew")
        ttk.Button(self.right_frame, text="Export NYSE Qualifying Stocks", command=self.export_nyse_qualifying_stocks).grid(row=3, column=0, pady=2, sticky="ew")

        ttk.Separator(self.right_frame, orient="horizontal").grid(row=4, column=0, sticky="ew", pady=5)

        ttk.Checkbutton(self.right_frame, text="Run NASDAQ Graham Screening", variable=self.nasdaq_screen_var).grid(row=5, column=0, pady=2, padx=2, sticky="w")
        ttk.Button(self.right_frame, text="Run NASDAQ Screening", command=self.run_nasdaq_screening).grid(row=6, column=0, pady=2, sticky="ew")
        ttk.Button(self.right_frame, text="Show NASDAQ Qualifying Stocks", command=self.display_nasdaq_qualifying_stocks).grid(row=7, column=0, pady=2, sticky="ew")
        ttk.Button(self.right_frame, text="Export NASDAQ Qualifying Stocks", command=self.export_nasdaq_qualifying_stocks).grid(row=8, column=0, pady=2, sticky="ew")

        # Add minimum criteria selector
        self.min_criteria_var = tk.IntVar(value=6)
        ttk.Label(self.right_frame, text="Min Graham Criteria:").grid(row=9, column=0, pady=2, sticky="w")
        self.min_criteria_menu = ttk.Combobox(self.right_frame, textvariable=self.min_criteria_var, values=[4, 5, 6], state="readonly")
        self.min_criteria_menu.grid(row=10, column=0, pady=2, sticky="ew")

        # Add sector filter selector below min criteria
        self.sector_filter_var = tk.StringVar(value="All")
        ttk.Label(self.right_frame, text="Sector Filter:").grid(row=11, column=0, pady=2, sticky="w")
        self.sector_menu = ttk.Combobox(
            self.right_frame, 
            textvariable=self.sector_filter_var, 
            values=[
                "All", "Energy", "Materials", "Industrials", "Consumer Discretionary", 
                "Consumer Staples", "Health Care", "Financials", "Information Technology", 
                "Communication Services", "Utilities", "Real Estate", "Unknown"
            ], 
            state="readonly"
        )
        self.sector_menu.grid(row=12, column=0, pady=2, sticky="ew")
        ttk.Checkbutton(self.right_frame, text="Screen Financials Separately", variable=self.financial_screen_var).grid(row=13, column=0, pady=2, padx=2, sticky="w")

        # Treeview and Tabs
        self.full_tree_frame = ttk.Frame(self.main_frame)
        self.full_tree_frame.grid(row=1, column=0, columnspan=3, sticky="nsew", padx=5, pady=5)
        self.full_tree_frame.grid_columnconfigure(0, weight=1)
        self.full_tree_frame.grid_rowconfigure(0, weight=3)
        self.full_tree_frame.grid_rowconfigure(1, weight=0)
        self.full_tree_frame.grid_rowconfigure(2, weight=1)

        self.tree = ttk.Treeview(self.full_tree_frame, columns=(1, 2, 3, 4, 5, 6, 7), show="tree headings", height=10)
        self.v_scrollbar = ttk.Scrollbar(self.full_tree_frame, orient="vertical", command=self.tree.yview)
        self.h_scrollbar = ttk.Scrollbar(self.full_tree_frame, orient="horizontal", command=self.tree.xview)
        self.tree.grid(row=0, column=0, sticky="nsew")
        self.v_scrollbar.grid(row=0, column=1, sticky="ns")
        self.h_scrollbar.grid(row=1, column=0, sticky="ew")
        self.tree.configure(yscrollcommand=self.v_scrollbar.set, xscrollcommand=self.h_scrollbar.set)

        self.tree.heading("#0", text="Ticker")
        self.tree.heading(1, text="Company Name")
        self.tree.heading(2, text="Sector")
        self.tree.heading(3, text="Score")
        self.tree.heading(4, text="Price ($)")
        self.tree.heading(5, text="Intrinsic Value ($)")
        self.tree.heading(6, text="Buy Price ($)")
        self.tree.heading(7, text="Sell Price ($)")
        self.tree.column("#0", width=80, anchor="center")
        self.tree.column(1, width=150, anchor="center")
        self.tree.column(2, width=80, anchor="center")
        self.tree.column(3, width=80, anchor="center")
        self.tree.column(4, width=80, anchor="center")
        self.tree.column(5, width=100, anchor="center")
        self.tree.column(6, width=80, anchor="center")
        self.tree.column(7, width=80, anchor="center")

        self.tree.tag_configure('highlight', background='lightgreen')

        for col in self.tree["columns"]:
            self.tree.heading(col, command=lambda c=col: self.sort_tree(self.tree["columns"].index(c)))

        self.data_frame = ttk.Frame(self.full_tree_frame)
        self.notebook = ttk.Notebook(self.data_frame)
        self.historical_frame = ttk.Frame(self.notebook)
        self.metrics_frame = ttk.Frame(self.notebook)
        self.notebook.add(self.historical_frame, text="Historical Data")
        self.notebook.add(self.metrics_frame, text="Metrics")
        self.data_frame.grid(row=2, column=0, columnspan=2, sticky="nsew")
        self.notebook.grid(row=0, column=0, sticky="nsew")
        self.data_frame.grid_columnconfigure(0, weight=1)
        self.data_frame.grid_rowconfigure(0, weight=1)

        self.historical_frame.grid_rowconfigure(0, weight=1)
        self.historical_frame.grid_columnconfigure(0, weight=1)
        self.metrics_frame.grid_rowconfigure(0, weight=1)
        self.metrics_frame.grid_columnconfigure(0, weight=1)

        self.historical_text = scrolledtext.ScrolledText(self.historical_frame, width=70, height=12)
        self.historical_text.grid(row=0, column=0, sticky="nsew")
        self.metrics_text = scrolledtext.ScrolledText(self.metrics_frame, width=70, height=12)
        self.metrics_text.grid(row=0, column=0, sticky="nsew")

        self.tree.bind("<<TreeviewSelect>>", self.update_tabs)

    def validate_tickers(self):
        tickers = self.parse_tickers(self.entry.get())
        valid = all(self.ticker_manager.is_valid_ticker(t) for t in tickers)
        self.entry.config(style="Invalid.TEntry" if not valid else "TEntry")
        # In __init__, add: style.configure("Invalid.TEntry", fieldbackground="pink")
        return valid

    def parse_tickers(self, tickers_input):
        if not tickers_input or not tickers_input.strip():
            return []
        tickers = [t.strip().upper() for t in tickers_input.split(',') if t.strip() and t.strip().isalnum() and len(t.strip()) <= 5]
        return list(set(tickers))

    def load_favorites(self):
        with FAVORITES_LOCK:
            if os.path.exists(FAVORITES_FILE):
                try:
                    with open(FAVORITES_FILE, 'r') as f:
                        favorites = json.load(f)
                        return {k: v if isinstance(v, list) else [] for k, v in favorites.items()}
                except json.JSONDecodeError:
                    analyze_logger.error(f"Corrupted favorites file {FAVORITES_FILE}")
                    return {}
            return {}

    def save_favorites(self):
        with FAVORITES_LOCK:
            try:
                if os.path.exists(FAVORITES_FILE):
                    shutil.copy(FAVORITES_FILE, FAVORITES_FILE + ".bak")
                with open(FAVORITES_FILE, 'w') as f:
                    json.dump(self.favorites, f, indent=4)
            except Exception as e:
                analyze_logger.error(f"Failed to save favorites: {str(e)}")
                messagebox.showerror("Error", f"Failed to save favorites: {str(e)}")

    def update_progress_animated(self, progress, tickers=None, passed_tickers=0, eta=None):
        self.progress_var.set(progress)
        if tickers is not None and isinstance(tickers, (list, tuple)):
            total_tickers = len(tickers)
            processed = int(progress / 100 * total_tickers)
            eta_text = f", ETA: {str(timedelta(seconds=int(eta)))[:-3]}" if eta is not None else ""
            self.progress_label.config(text=f"Progress: {progress:.1f}% ({processed}/{total_tickers} stocks processed, {passed_tickers} passed{eta_text})")
        else:
            eta_text = f", ETA: {str(timedelta(seconds=int(eta)))[:-3]}" if eta is not None else ""
            self.progress_label.config(text=f"Progress: {progress:.1f}% (Screening, {passed_tickers} passed{eta_text})")
        self.root.update_idletasks()

    def update_cache_usage(self, cache_hits, total):
        if total > 0:
            cache_percent = (cache_hits / total) * 100
            self.cache_var.set(cache_percent)
            self.cache_label.config(text=f"Cache Usage: {cache_percent:.1f}% ({cache_hits} cached, {total - cache_hits} fresh)")
        else:
            self.cache_var.set(0)
            self.cache_label.config(text="Cache Usage: 0% (0 cached, 0 fresh)")
        self.root.update_idletasks()

    def update_rate_limit(self, message):
        self.rate_limit_label.config(text=message)
        self.root.update_idletasks()

    def refresh_favorites_dropdown(self, selected_list=None):
        self.favorites = self.load_favorites()
        analyze_logger.debug(f"Refreshed favorites dropdown with keys: {list(self.favorites.keys())}")
        self.favorite_menu['values'] = list(self.favorites.keys())
        if selected_list and selected_list in self.favorites:
            self.favorite_var.set(selected_list)
            analyze_logger.debug(f"Set dropdown to: {selected_list}")

    def cancel_screening(self):
        if messagebox.askyesno("Confirm Cancel", "Are you sure you want to cancel the screening?"):
            self.cancel_event.set()
            self.status_label.config(text="Cancelling screening...")

    # Screening Logic
    def run_screening(self, exchange: str, screen_func) -> None:
        if self.screening_active:
            messagebox.showwarning("Warning", f"A {exchange} screening is already in progress.")
            return
        if (exchange == "NYSE" and not self.nyse_screen_var.get()) or (exchange == "NASDAQ" and not self.nasdaq_screen_var.get()):
            return
        if self.nyse_screen_var.get() and self.nasdaq_screen_var.get():
            messagebox.showwarning("Warning", "Cannot run NYSE and NASDAQ screening simultaneously.")
            self.nyse_screen_var.set(False)
            self.nasdaq_screen_var.set(False)
            return

        file_path = NYSE_LIST_FILE if exchange == "NYSE" else NASDAQ_LIST_FILE
        if not os.path.exists(file_path):
            messagebox.showerror("Error", f"{exchange} list file missing: {file_path}")
            return

        file_age = time.time() - os.path.getmtime(file_path)
        if file_age > 365 * 24 * 60 * 60:
            messagebox.showwarning("Old Data", f"{exchange} list file is over 365 days old.")

        conn, cursor = get_stocks_connection()
        try:
            cursor.execute("SELECT MAX(timestamp) FROM stocks")
            max_timestamp = cursor.fetchone()[0]
            if max_timestamp and time.time() - max_timestamp > 365 * 24 * 60 * 60:
                messagebox.showwarning("Old Data", "The cached data is over a year old. Consider refreshing the data.")
        except sqlite3.Error as e:
            analyze_logger.error(f"Error checking data age: {str(e)}")
        finally:
            conn.close()

        self.progress_var.set(0)
        self.cache_var.set(0)
        self.progress_label.config(text=f"Progress: 0% (0/0 stocks processed, 0 passed)")
        self.cache_label.config(text="Cache Usage: 0% (0 cached, 0 fresh)")
        self.rate_limit_label.config(text="")
        self.status_label.config(text=f"Running {exchange} screening...")
        self.root.update()
        self.cancel_event.clear()
        self.screening_active = True

        min_criteria = self.min_criteria_var.get()
        sector_filter = self.sector_filter_var.get() if self.sector_filter_var.get() != "All" else None  # New: Get sector
        separate_financials = self.financial_screen_var.get()  # Added: Pass checkbox value

        async def screening_task():
            try:
                await self.ticker_manager.initialize()
                qualifying_stocks, common_scores, exchanges, error_tickers, financial_qualifying_stocks = await screen_func(  # Updated return unpacking
                    exchange=exchange,  # New param
                    batch_size=50,
                    cancel_event=self.cancel_event,
                    root=self.root,
                    update_progress_animated=self.update_progress_animated,
                    refresh_favorites_dropdown=self.refresh_favorites_dropdown,
                    ticker_manager=self.ticker_manager,
                    update_rate_limit=self.update_rate_limit,
                    min_criteria=min_criteria,
                    sector_filter=sector_filter,  # New: Pass sector
                    separate_financials=separate_financials  # Added
                )
                screening_logger.info(f"Qualifying stocks for {exchange}: {qualifying_stocks}")
                if not self.cancel_event.is_set():
                    if qualifying_stocks:
                        list_name = await save_qualifying_stocks_to_favorites(qualifying_stocks, exchange)
                        if list_name:
                            self.root.after(0, lambda: self.refresh_favorites_dropdown(list_name))
                        else:
                            screening_logger.error(f"Failed to save qualifying stocks for {exchange}")
                    else:
                        screening_logger.info(f"No qualifying stocks found for {exchange}")
                    conn, cursor = get_stocks_connection()
                    try:
                        processed_tickers = cursor.execute(
                            "SELECT COUNT(*) FROM screening_progress WHERE exchange=? AND status='completed'",
                            (exchange,)
                        ).fetchone()[0]
                        total_tickers = processed_tickers + len(error_tickers)
                        ticker_list = list(self.ticker_manager.get_tickers(exchange))
                        cache_hits = sum(
                            1 for t in ticker_list
                            if get_stock_data_from_db(t) and time.time() - get_stock_data_from_db(t)['timestamp'] < CACHE_EXPIRY
                        )
                        self.root.after(0, lambda: self.update_cache_usage(cache_hits, total_tickers))
                        summary = f"Completed {exchange} screening.\nProcessed {total_tickers} stocks,\nFound {len(qualifying_stocks)} qualifiers,\n{len(error_tickers)} errors"
                        if separate_financials:
                            summary += f"\nFinancial qualifiers: {len(financial_qualifying_stocks)}"
                        if error_tickers:
                            summary += f"\nError tickers: {', '.join(error_tickers)}"
                        self.root.after(0, lambda: messagebox.showinfo("Screening Complete", summary))
                    except sqlite3.Error as e:
                        screening_logger.error(f"Database error in screening summary: {str(e)}")
                    finally:
                        conn.close()
            except Exception as e:
                self.root.after(0, lambda: messagebox.showerror("Error", f"Screening failed: {str(e)}"))
                screening_logger.error(f"Screening task failed for {exchange}: {str(e)}", exc_info=True)
            finally:
                self.root.after(0, lambda: self.progress_label.config(text=f"Progress: 100% (Screening Complete - {exchange})"))
                self.root.after(0, lambda: self.status_label.config(text=""))
                setattr(self, f"{exchange.lower()}_screen_var", tk.BooleanVar(value=False))
                self.screening_active = False

        self.task_queue.put(screening_task())

    def run_nyse_screening(self) -> None:
        self.run_screening("NYSE", screen_exchange_graham_stocks)  # Updated to use refactored func

    def run_nasdaq_screening(self) -> None:
        self.run_screening("NASDAQ", screen_exchange_graham_stocks)  # Updated to use refactored func

    # Data Fetching and Analysis
    async def fetch_company_name(self, ticker):
        try:
            stock = yf.Ticker(ticker)
            return stock.info.get('longName', 'Unknown')
        except Exception as e:
            analyze_logger.error(f"Error fetching company name for {ticker}: {str(e)}")
            return 'Unknown'

    def format_float(self, value, precision=2):
        if pd.isna(value) or not isinstance(value, (int, float)):
            return "N/A"
        return f"{value:.{precision}f}"

    async def fetch_cached_data(self, ticker, exchange="Unknown"):
        nyse_file_hash = get_file_hash(NYSE_LIST_FILE)
        nasdaq_file_hash = get_file_hash(NASDAQ_LIST_FILE)

        conn, cursor = get_stocks_connection()
        try:
            cursor.execute("SELECT file_hash, exchange, timestamp FROM screening_progress WHERE ticker=?", (ticker,))
            row = cursor.fetchone()
            if row:
                columns = [desc[0] for desc in cursor.description]
                progress_dict = dict(zip(columns, row))
                stored_hash = progress_dict['file_hash']
                stored_exchange = progress_dict['exchange']
                timestamp_str = progress_dict['timestamp']
                timestamp = datetime.strptime(timestamp_str, '%Y-%m-%d %H:%M:%S').timestamp() if timestamp_str else 0
                if ((stored_exchange == "NYSE" and stored_hash == nyse_file_hash) or 
                    (stored_exchange == "NASDAQ" and stored_hash == nasdaq_file_hash)) and \
                (time.time() - timestamp < CACHE_EXPIRY):
                    cursor.execute("SELECT * FROM stocks WHERE ticker=?", (ticker,))
                    stock_row = cursor.fetchone()
                    if stock_row:
                        columns = [desc[0] for desc in cursor.description]
                        stock_dict = dict(zip(columns, stock_row))
                        analyze_logger.info(f"Database cache hit for {ticker}")
                        stock = yf.Ticker(ticker)
                        info = stock.info
                        price = info.get('regularMarketPrice', info.get('previousClose', None))
                        if price is None:
                            analyze_logger.error(f"No price data for {ticker} from YFinance")
                            return None
                        roe_list = [float(x) if x.strip() else 0.0 for x in stock_dict['roe'].split(",")] if stock_dict['roe'] else []
                        rotc_list = [float(x) if x.strip() else 0.0 for x in stock_dict['rotc'].split(",")] if stock_dict['rotc'] else []
                        eps_list = [float(x) if x.strip() else 0.0 for x in stock_dict['eps'].split(",")] if stock_dict['eps'] else []
                        div_list = [float(x) if x.strip() else 0.0 for x in stock_dict['dividend'].split(",")] if stock_dict['dividend'] else []
                        balance_data = json.loads(stock_dict['balance_data']) if stock_dict['balance_data'] else []
                        years = [int(y) for y in stock_dict['years'].split(",")] if stock_dict.get('years') else []

                        company_name = stock_dict['company_name']
                        debt_to_equity = stock_dict['debt_to_equity']
                        eps_ttm = stock_dict['eps_ttm']
                        book_value_per_share = stock_dict['book_value_per_share']
                        sector = stock_dict.get('sector', 'Unknown')

                        cached_stock_data = {
                            "ticker": stock_dict['ticker'],
                            "sector": sector,
                            "eps_ttm": eps_ttm,
                            "eps_list": eps_list,
                            "eps_cagr": stock_dict.get('eps_cagr', 0.0)
                        }
                        intrinsic_value = await calculate_graham_value(eps_ttm, cached_stock_data) if eps_ttm is not None and eps_ttm > 0 else float('nan')
                        margin_of_safety = self.margin_of_safety_var.get() / 100
                        expected_return = self.expected_return_var.get() / 100
                        buy_price = intrinsic_value * (1 - margin_of_safety) if not pd.isna(intrinsic_value) else float('nan')
                        sell_price = intrinsic_value * (1 + expected_return) if not pd.isna(intrinsic_value) else float('nan')
                        latest_revenue = stock_dict['latest_revenue']
                        key_metrics_data = json.loads(stock_dict.get('key_metrics_data', '[]'))
                        graham_score = calculate_graham_score_8(ticker, price, None, None, debt_to_equity, eps_list, div_list, {}, balance_data, key_metrics_data, stock_dict['available_data_years'], latest_revenue, sector)

                        result = {
                            "ticker": stock_dict['ticker'],
                            "sector": sector,
                            "company_name": company_name,
                            "price": price,
                            "intrinsic_value": intrinsic_value,
                            "buy_price": buy_price,
                            "sell_price": sell_price,
                            "graham_score": graham_score,
                            "years": years,
                            "roe_list": roe_list,
                            "rotc_list": rotc_list,
                            "eps_list": eps_list,
                            "div_list": div_list,
                            "balance_data": balance_data,
                            "latest_revenue": latest_revenue,
                            "available_data_years": stock_dict['available_data_years'],
                            "eps_ttm": eps_ttm,
                            "book_value_per_share": book_value_per_share,
                            "debt_to_equity": debt_to_equity,
                            "eps_cagr": stock_dict.get('eps_cagr', 0.0)
                        }
                        with self.ticker_cache_lock:
                            self.ticker_cache[ticker] = result
                        analyze_logger.debug(f"Cached data retrieved for {ticker}: Score={graham_score}/8 with {stock_dict['available_data_years']} years")
                        return result
            analyze_logger.info(f"Cache miss for {ticker}: Data stale or missing")
            return None
        except sqlite3.Error as e:
            analyze_logger.error(f"Database error in fetch_cached_data for {ticker}: {str(e)}")
            return None
        finally:
            conn.close()

    async def determine_exchange(self, ticker):
        await self.ticker_manager.initialize()
        nyse_tickers = self.ticker_manager.get_tickers("NYSE")
        nasdaq_tickers = self.ticker_manager.get_tickers("NASDAQ")
        if ticker in nyse_tickers:
            return "NYSE"
        elif ticker in nasdaq_tickers:
            return "NASDAQ"
        return "Unknown"

    async def analyze_multiple_stocks_async(self, tickers):
        analyze_logger.info(f"Starting analysis for tickers: {tickers}")

        nyse_invalid_file = os.path.join(USER_DATA_DIR, "NYSE Invalid Tickers.txt")
        nasdaq_invalid_file = os.path.join(USER_DATA_DIR, "NASDAQ Invalid Tickers.txt")
        invalid_tickers = set()
        for invalid_file in [nyse_invalid_file, nasdaq_invalid_file]:
            if os.path.exists(invalid_file):
                with open(invalid_file, 'r') as f:
                    invalid_tickers.update(f.read().splitlines())
        valid_tickers = [t for t in tickers if t not in invalid_tickers]
        if len(valid_tickers) < len(tickers):
            excluded = set(tickers) - set(valid_tickers)
            analyze_logger.info(f"Excluded {len(excluded)} invalid tickers: {excluded}")
            self.root.after(0, lambda: messagebox.showinfo("Excluded Tickers", f"Excluded {len(excluded)} invalid tickers: {', '.join(excluded)}"))
        tickers = valid_tickers

        if not tickers:
            analyze_logger.warning("No valid tickers provided for analysis")
            messagebox.showwarning("No Tickers", "No valid tickers to analyze.")
            return

        conn, cursor = get_stocks_connection()
        try:
            cursor.execute("SELECT MAX(timestamp) FROM stocks")
            max_timestamp = cursor.fetchone()[0]
            if max_timestamp and time.time() - max_timestamp > 365 * 24 * 60 * 60:
                analyze_logger.warning("Cached data is over a year old")
                messagebox.showwarning("Old Data", "The cached data is over a year old. Consider refreshing the data.")
            else:
                analyze_logger.info(f"Most recent data timestamp: {max_timestamp}")
        except sqlite3.Error as e:
            analyze_logger.error(f"Error checking data age: {str(e)}")
        finally:
            conn.close()

        self.progress_var.set(0)
        self.cache_var.set(0)
        self.progress_label.config(text=f"Progress: 0% (0/{len(tickers)} stocks processed, 0 passed)")
        self.cache_label.config(text="Cache Usage: 0% (0 cached, 0 fresh)")
        self.rate_limit_label.config(text="")
        self.status_label.config(text="Analyzing stocks...")
        self.root.update()

        analyze_logger.info(f"Fetching data for {len(tickers)} tickers with margin_of_safety={self.margin_of_safety_var.get()/100}, expected_return={self.expected_return_var.get()/100}")
        results, error_tickers, cache_hits = await fetch_batch_data(
            tickers,
            screening_mode=False,
            expected_return=self.expected_return_var.get() / 100,
            margin_of_safety=self.margin_of_safety_var.get() / 100,
            ticker_manager=self.ticker_manager,
            update_rate_limit=self.update_rate_limit
        )

        valid_results = [r for r in results if 'error' not in r]
        passed_tickers = sum(1 for r in valid_results if r.get('graham_score', 0) >= 5 and r.get('available_data_years', 0) >= 10)

        analyze_logger.info(f"Analysis complete: {len(valid_results)} valid results, {len(error_tickers)} errors, {cache_hits} cache hits")
        if error_tickers:
            analyze_logger.warning(f"Error tickers: {error_tickers}")
            error_summary = "\n".join(error_tickers)
            self.root.after(0, lambda: messagebox.showwarning("Analysis Errors", f"Errors occurred for the following tickers:\n{error_summary}"))

        analyze_logger.debug("Populating treeview with analysis results")
        for item in self.tree.get_children():
            self.tree.delete(item)

        for result in valid_results:
            years_used = result.get('available_data_years', 0)
            analyze_logger.debug(f"Processing result for {result['ticker']}: Graham Score={result['graham_score']}, Years={years_used}")
            if years_used < 10:
                analyze_logger.warning(f"{result['ticker']} has only {years_used} years of data")
                self.root.after(0, lambda t=result['ticker'], y=years_used: messagebox.showwarning("Insufficient Data", f"{t} has only {y} years of data. Results may be incomplete."))
            warning = f" (based on {years_used} years)" if years_used < 10 else ""

            price = result['price']
            buy_price = result['buy_price']
            tags = ('highlight',) if price <= buy_price and not pd.isna(price) and not pd.isna(buy_price) else ()
            self.tree.insert("", "end", text=result['ticker'], values=(
                result['company_name'],
                result['sector'],
                f"{result['graham_score']}/8{warning}",
                f"${self.format_float(result['price'])}",
                f"${self.format_float(result['intrinsic_value'])}",
                f"${self.format_float(result['buy_price'])}",
                f"${self.format_float(result['sell_price'])}"
            ), tags=tags)

            # Cache the result in ticker_cache
            with self.ticker_cache_lock:
                self.ticker_cache[result['ticker']] = result

        analyze_logger.info(f"Cache usage: {cache_hits} hits out of {len(tickers)} total tickers")
        self.root.after(0, lambda: self.update_cache_usage(cache_hits, len(tickers)))
        self.root.after(0, lambda: self.update_progress_animated(100, tickers, passed_tickers))
        self.root.after(0, lambda: self.status_label.config(text=""))
        analyze_logger.info("Analysis fully completed and UI updated")

    def analyze_multiple_stocks(self, tickers_input=None):
        analyze_logger.info("Analyze Stocks button clicked")
        if self.analysis_lock.acquire(timeout=5):
            try:
                if tickers_input is None:
                    tickers_input = self.entry.get()
                tickers = self.parse_tickers(tickers_input)
                analyze_logger.info(f"Parsed tickers from input '{tickers_input}': {tickers}")
                if not tickers:
                    analyze_logger.warning("No tickers parsed from input")
                self.task_queue.put(self.analyze_multiple_stocks_async(tickers))
            finally:
                self.analysis_lock.release()
        else:
            analyze_logger.error("Failed to acquire analysis lock within 5 seconds")
            messagebox.showerror("Error", "Unable to start analysis: Lock timeout.")

    async def display_results_in_tree(self, tickers, scores, exchanges, source):
        for item in self.tree.get_children():
            self.tree.delete(item)

        async def fetch_company_names(tickers_list):
            tasks = [self.fetch_company_name(ticker) for ticker in tickers_list]
            return await asyncio.gather(*tasks)

        results = []
        for ticker in tickers:
            with self.ticker_cache_lock:
                if ticker in self.ticker_cache:
                    results.append(self.ticker_cache[ticker])
                    continue

            cached_result = await self.fetch_cached_data(ticker, source)
            if cached_result:
                results.append(cached_result)
                with self.ticker_cache_lock:
                    self.ticker_cache[ticker] = cached_result
                continue

            result = await fetch_batch_data(
                [ticker], 
                screening_mode=False, 
                expected_return=self.expected_return_var.get() / 100,
                margin_of_safety=self.margin_of_safety_var.get() / 100,
                exchange=source, 
                ticker_manager=self.ticker_manager,
                update_rate_limit=self.update_rate_limit
            )
            if result and isinstance(result[0], dict) and "error" not in result[0]:
                result = result[0]
                with self.ticker_cache_lock:
                    self.ticker_cache[ticker] = result
                results.append(result)
            else:
                results.append({"ticker": ticker, "exchange": source, "error": "Failed to fetch data"})

        valid_results = [r for r in results if 'error' not in r]
        company_names = await fetch_company_names(tickers)

        display_list = []
        for i, ticker in enumerate(tickers):
            if i < len(valid_results):
                result = valid_results[i]
                company_name = company_names[i] if i < len(company_names) else 'Unknown'
                display_list.append((ticker, company_name, result['exchange'], scores[i], result))
            else:
                display_list.append((ticker, 'Unknown', exchanges[i] if i < len(exchanges) else source, scores[i], {}))

        display_list.sort(key=lambda x: x[0])

        for ticker, company_name, exchange, common_score, result in display_list:
            price = result.get('price', 0)
            buy_price = result.get('buy_price', float('nan'))
            tags = ('highlight',) if price <= buy_price and not pd.isna(price) and not pd.isna(buy_price) else ()
            self.tree.insert("", "end", text=ticker, values=(
                company_name,
                exchange,
                f"{common_score}/6",
                f"${self.format_float(result.get('price', 0))}",
                f"${self.format_float(result.get('intrinsic_value', float('nan')))}",
                f"${self.format_float(result.get('buy_price', float('nan')))}",
                f"${self.format_float(result.get('sell_price', float('nan')))}"
            ), tags=tags)

        self.sort_tree(0)

    def update_tabs(self, event):
        selected = self.tree.selection()
        if not selected:
            return
        ticker = self.tree.item(selected[0], "text")

        with self.ticker_cache_lock:
            result = self.ticker_cache.get(ticker, None)

        if not result:
            analyze_logger.warning(f"No cached data found for {ticker}, fetching fresh data")
            async def fetch_fresh():
                fresh_result = await fetch_stock_data(
                    ticker,
                    expected_return=self.expected_return_var.get() / 100,
                    margin_of_safety=self.margin_of_safety_var.get() / 100,
                    ticker_manager=self.ticker_manager
                )
                with self.ticker_cache_lock:
                    self.ticker_cache[ticker] = fresh_result
                return fresh_result
            result = asyncio.run(fetch_fresh())

        analyze_logger.info(f"Updating tabs for {ticker} with data keys: {list(result.keys())}")

        self.root.after(0, lambda: self.historical_text.delete('1.0', tk.END))
        self.root.after(0, lambda: self.metrics_text.delete('1.0', tk.END))

        years = result.get('years', [])
        roe_list = result.get('roe_list', [])
        rotc_list = result.get('rotc_list', [])
        eps_list = result.get('eps_list', [])
        div_list = result.get('div_list', [])

        analyze_logger.debug(f"Historical data for {ticker}: years={len(years)}, roe={len(roe_list)}, rotc={len(rotc_list)}, eps={len(eps_list)}, dividend={len(div_list)}")
        analyze_logger.debug(f"Years: {years}")
        analyze_logger.debug(f"ROE: {roe_list}")
        analyze_logger.debug(f"ROTC: {rotc_list}")
        analyze_logger.debug(f"EPS: {eps_list}")
        analyze_logger.debug(f"Dividend: {div_list}")

        if not years:
            self.root.after(0, lambda: self.historical_text.insert(tk.END, f"No historical data available for {ticker}\n"))
        elif len(years) == len(roe_list) == len(rotc_list) == len(eps_list) == len(div_list):
            header = f"{len(years)}-Year Historical Data for {ticker}:\nYear\tROE (%)\tROTC (%)\tEPS ($)\tDividend ($)\n"
            self.root.after(0, lambda: self.historical_text.insert(tk.END, header))
            for j in range(len(years)):
                line = f"{years[j]}\t{roe_list[j]:.2f}\t{rotc_list[j]:.2f}\t{eps_list[j]:.2f}\t{div_list[j]:.2f}\n"
                self.root.after(0, lambda l=line: self.historical_text.insert(tk.END, l))
        else:
            message = f"Incomplete Historical Data for {ticker} (Years: {len(years)}, ROE: {len(roe_list)}, ROTC: {len(rotc_list)}, EPS: {len(eps_list)}, Dividend: {len(div_list)})\n"
            self.root.after(0, lambda m=message: self.historical_text.insert(tk.END, m))

        self.root.after(0, lambda: self.metrics_text.insert(tk.END, f"Graham Criteria Results for {ticker} (Score: {result['graham_score']}/8 with {result['available_data_years']} years):\n"))
        self.root.after(0, lambda: self.metrics_text.insert(tk.END, f"Sector: {result['sector']} (Growth Rate: {get_sector_growth_rate(result['sector'])}%)\n\n"))

        async def update_metrics():
            stock_data = get_stock_data_from_db(ticker)
            if not stock_data:
                self.safe_insert(self.metrics_text, f"No data available for {ticker}\n")
                return

            eps_list = stock_data['eps_list']
            div_list = stock_data['div_list']
            balance_data = stock_data['balance_data']
            latest_revenue = stock_data['latest_revenue']
            debt_to_equity = stock_data['debt_to_equity']
            available_data_years = stock_data['available_data_years']
            sector = stock_data['sector']
            eps_ttm = stock_data['eps_ttm']
            book_value_per_share = stock_data['book_value_per_share']
            price = result.get('price', None)

            revenue_passed = latest_revenue >= 500_000_000
            self.safe_insert(self.metrics_text, f"1. Revenue >= $500M: {'Yes' if revenue_passed else 'No'} (${latest_revenue / 1e6:.2f}M)\n")

            if balance_data and 'totalCurrentAssets' in balance_data[0] and 'totalCurrentLiabilities' in balance_data[0]:
                current_assets = float(balance_data[0].get('totalCurrentAssets', 0))
                current_liabilities = float(balance_data[0].get('totalCurrentLiabilities', 1))
                current_ratio = current_assets / current_liabilities if current_liabilities > 0 else 0
                current_passed = current_ratio > 2
                self.safe_insert(self.metrics_text, f"2. Current Ratio > 2: {'Yes' if current_passed else 'No'} ({current_ratio:.2f})\n")
            else:
                self.safe_insert(self.metrics_text, "2. Current Ratio > 2: No (Missing data)\n")

            negative_eps_count = sum(1 for eps in eps_list if eps <= 0)
            stability_passed = negative_eps_count == 0
            self.safe_insert(self.metrics_text, f"3. All Positive EPS: {'Yes' if stability_passed else 'No'} (Negative EPS years: {negative_eps_count})\n")

            dividend_passed = all(div > 0 for div in div_list) if div_list else False
            div_display = f" (Dividends: {', '.join(f'{d:.2f}' for d in div_list)})" if div_list else ""
            self.safe_insert(self.metrics_text, f"4. Uninterrupted Dividends: {'Yes' if dividend_passed else 'No'}{div_display}\n")

            if len(eps_list) >= 2 and eps_list[0] > 0 and eps_list[-1] > 0:
                cagr = calculate_cagr(eps_list[0], eps_list[-1], available_data_years - 1)
                growth_passed = cagr > 0.03
                self.safe_insert(self.metrics_text, f"5. EPS CAGR > 3%: {'Yes' if growth_passed else 'No'} ({cagr:.2%})\n")
            else:
                eps_display = f" (EPS: {', '.join(f'{e:.2f}' for e in eps_list)})" if eps_list else ""
                self.safe_insert(self.metrics_text, f"5. EPS CAGR > 3%: No (Insufficient or invalid data{eps_display})\n")

            debt_passed = debt_to_equity is not None and debt_to_equity < 2
            self.safe_insert(self.metrics_text, f"6. Debt-to-Equity < 2: {'Yes' if debt_passed else 'No'} ({self.format_float(debt_to_equity)})\n")

            pe_ratio = price / eps_ttm if price and eps_ttm and eps_ttm > 0 else None
            pe_pass = pe_ratio is not None and pe_ratio <= 15
            self.safe_insert(self.metrics_text, f"7. P/E Ratio <= 15: {'Yes' if pe_pass else 'No'} ({self.format_float(pe_ratio)})\n")

            pb_ratio = price / book_value_per_share if price and book_value_per_share and book_value_per_share > 0 else None
            pb_pass = pb_ratio is not None and pb_ratio <= 1.5
            self.safe_insert(self.metrics_text, f"8. P/B Ratio <= 1.5: {'Yes' if pb_pass else 'No'} ({self.format_float(pb_ratio)})\n")

            analyze_logger.info(f"Metrics updated for {ticker} using cached data")

        self.task_queue.put(update_metrics())

    def manage_favorites(self):
        manage_window = tk.Toplevel(self.root)
        manage_window.title("Manage Favorites")
        manage_window.geometry("400x300")

        favorites_listbox = tk.Listbox(manage_window, height=15)
        favorites_listbox.pack(fill="both", expand=True, padx=5, pady=5)
        for name in self.favorites.keys():
            favorites_listbox.insert(tk.END, name)

        def delete_favorite():
            selected = favorites_listbox.curselection()
            if selected:
                name = favorites_listbox.get(selected[0])
                del self.favorites[name]
                self.save_favorites()
                favorites_listbox.delete(selected[0])
                self.refresh_favorites_dropdown()

        ttk.Button(manage_window, text="Delete Selected", command=delete_favorite).pack(pady=5)
        ttk.Button(manage_window, text="Close", command=manage_window.destroy).pack(pady=5)

    def sort_tree(self, col):
        items = [(self.tree.set(k, col), k) for k in self.tree.get_children()]
        items.sort()
        for index, (val, k) in enumerate(items):
            self.tree.move(k, '', index)

    def display_nyse_qualifying_stocks(self):
        """Display NYSE qualifying stocks in the treeview with company names and scores out of 6."""
        conn, cursor = get_stocks_connection()
        try:
            cursor.execute("""
                SELECT g.ticker, s.company_name, g.sector, g.common_score
                FROM graham_qualifiers g
                JOIN stocks s ON g.ticker = s.ticker
                WHERE g.exchange = 'NYSE' AND g.min_criteria = ?
                ORDER BY g.common_score DESC, g.ticker ASC
            """, (self.min_criteria_var.get(),))
            results = cursor.fetchall()

            # Clear tree
            for item in self.tree.get_children():
                self.tree.delete(item)

            if not results:
                self.tree.insert("", "end", text="No qualifying stocks found",
                                values=("", "", "", "", "", "", ""))
                analyze_logger.info(f"No NYSE qualifiers found for min_criteria={self.min_criteria_var.get()}")
                return

            # Populate treeview
            for ticker, company_name, sector, common_score in results:
                display_name = company_name if company_name else ticker
                self.tree.insert("", "end", text=ticker, values=(
                    display_name, sector, f"{common_score}/6", "", "", "", ""
                ), tags=('financial',) if sector == 'Financials' else ())
                analyze_logger.debug(f"Displayed NYSE {ticker}: {display_name} (Sector: {sector}, Score: {common_score}/6)")

            self.tree.tag_configure('financial', foreground='blue')  # Visual distinction for financials
            self.sort_tree(0)
            analyze_logger.info(f"Displayed {len(results)} NYSE qualifying stocks")
        except sqlite3.Error as e:
            error_msg = f"Database error displaying NYSE qualifiers: {str(e)}"
            analyze_logger.error(error_msg)
            messagebox.showerror("Database Error", error_msg)
        finally:
            conn.close()

    def display_nasdaq_qualifying_stocks(self):
        """Display NASDAQ qualifying stocks in the treeview with company names and scores out of 6."""
        conn, cursor = get_stocks_connection()
        try:
            cursor.execute("""
                SELECT g.ticker, s.company_name, g.sector, g.common_score
                FROM graham_qualifiers g
                JOIN stocks s ON g.ticker = s.ticker
                WHERE g.exchange = 'NASDAQ' AND g.min_criteria = ?
                ORDER BY g.common_score DESC, g.ticker ASC
            """, (self.min_criteria_var.get(),))
            results = cursor.fetchall()

            # Clear tree
            for item in self.tree.get_children():
                self.tree.delete(item)

            if not results:
                self.tree.insert("", "end", text="No qualifying stocks found",
                                values=("", "", "", "", "", "", ""))
                analyze_logger.info(f"No NASDAQ qualifiers found for min_criteria={self.min_criteria_var.get()}")
                return

            # Populate treeview
            for ticker, company_name, sector, common_score in results:
                display_name = company_name if company_name else ticker
                self.tree.insert("", "end", text=ticker, values=(
                    display_name, sector, f"{common_score}/6", "", "", "", ""
                ), tags=('financial',) if sector == 'Financials' else ())
                analyze_logger.debug(f"Displayed NASDAQ {ticker}: {display_name} (Sector: {sector}, Score: {common_score}/6)")

            self.tree.tag_configure('financial', foreground='blue')  # Visual distinction for financials
            self.sort_tree(0)
            analyze_logger.info(f"Displayed {len(results)} NASDAQ qualifying stocks")
        except sqlite3.Error as e:
            error_msg = f"Database error displaying NASDAQ qualifiers: {str(e)}"
            analyze_logger.error(error_msg)
            messagebox.showerror("Database Error", error_msg)
        finally:
            conn.close()

    def fetch_current_prices(self, tickers):
        """Fetch current prices for a list of tickers using yfinance."""
        try:
            data = yf.download(tickers, period="1d")['Close']
            if len(tickers) == 1:
                return {tickers[0]: data.iloc[-1]}
            else:
                return data.iloc[-1].to_dict()
        except Exception as e:
            analyze_logger.error(f"Error fetching prices: {e}")
            return {ticker: None for ticker in tickers}

    def calculate_intrinsic_value(self, stock_dict):
        """Calculate Graham intrinsic value using EPS and EPS CAGR."""
        eps = stock_dict.get('eps_ttm')
        eps_cagr = stock_dict.get('eps_cagr', 0.0)
        if not eps or eps <= 0:
            return float('nan')
        aaa_yield = get_aaa_yield(FRED_API_KEY)
        if aaa_yield <= 0:
            return float('nan')
        g = eps_cagr * 100  # Convert decimal to percentage
        earnings_multiplier = min(8.5 + 2 * g, 20)  # Cap at 20
        normalization_factor = 4.4
        value = (eps * earnings_multiplier * normalization_factor) / (100 * aaa_yield)
        return value

    def setup_start_here_sheet(self, start_sheet):
        # Set tab color to yellow
        start_sheet.sheet_properties.tabColor = "FFFF00"

        # Set column widths to 10 for columns A to H
        for col in range(1, 9):
            start_sheet.column_dimensions[get_column_letter(col)].width = 10

        # Row 1: Title
        start_sheet['A1'] = "DIRECTIONS, GENERAL NOTES, CALCULATIONS AND CONSTANTS FOR FOLLOWING SPREADSHEETS"
        start_sheet['A1'].font = Font(bold=True, size=16)
        start_sheet['A1'].alignment = Alignment(horizontal='left', vertical='center')
        start_sheet.row_dimensions[1].height = 30

        # Rows 3 to 8: Instructions
        instructions = [
            "Go to NYSE Index below and select an alphabet Letter. Click on a company name hyperlink you'd like to analyze, then click on the company name hyperlink again. This will open the Morningstar quote page for the respective company.",
            "On the Morningstar quote page scroll down to 'financials' and click it. At the bottom of the table click 'oling Financials Data.' Click the 'balance sheet' link at the top. Record 'Total Assets' and 'Total Liabilities.'",
            "At the top of the 'oling Financials Data' page, select the 'Key Ratios' tab. From the 'Financials' section, record the 'Net Income,' 'Earnings Per Share,' 'Dividends,' 'Shares,' and 'Working Capital' for the most recent year.",
            "Staying on the 'Key Ratios' tab in the 'Key Ratios' section, record 'Return on Equity' and 'Return on Invested Capital' for the last 10 years.",
            "Finally, remaining on the 'Key Ratios' tab, in the 'Key Ratios' section, select the 'Financial Health' tab. Scroll down to 'Long Term Debt' and enter the most recent annual number in the spreadsheet.",
            "The spreadsheet will calculate the rest. Click save. You're done! Good job and may we get rich together!!"
        ]
        for i, text in enumerate(instructions, start=3):
            start_sheet[f'A{i}'] = text
            start_sheet[f'A{i}'].alignment = Alignment(horizontal='left', vertical='center')

        # Rows 11 to 20: Tips
        tips = [
            "Consistently high Return on Equity shows durability; ROE = Net Income / Book Value (should be better than 12%).",
            "Consistently high Return on Total Capital shows durability: ROTC = Net Earnings / Total Capital (should be better than 12% except for banks. Bank Return on Total Assets should be better than 1-1.5%).",
            "Consistently high Earnings Per Share (EPS) shows durable competitive advantage: Earnings Per Share = Total Net Earnings / Number of Shares Outstanding. Should be strong and show an upward trend. Also look for an upward trend then a sharp drop. Sometimes this is a sign of a one-time oops that the market overreacts to.",
            "Durable companies should have Long Term Debt of no more than 5 times current net earnings.",
            "The product the company sells should be something that people use every day but wears out quickly causing repurchase.",
            "Durable companies are not controlled by labor unions try to avoid them.",
            "Avoid companies that cannot sustain the price of their product in a down economy or a rise in inflation (i.e., airlines), good example - Coca-Cola.",
            "Avoid companies that have to reinvest their earnings in operational costs (i.e., General Motors, Bethlehem Steel), good example - H&R Block, Wrigleys (same products forever with very little change).",
            "Look for companies that buy back shares of their own stock. This shows they have excess cash and want to pay down debt.",
            "Look for a company that uses its earnings to increase overall market value (i.e., Berkshire Hathaway). Look at a ten-year spread on the company's share price and its book value. If these values do not show a significant increase over the last 10 years, it is not using its earnings to increase its worth."
        ]
        for i, tip in enumerate(tips, start=11):
            start_sheet[f'A{i}'] = tip
            start_sheet[f'A{i}'].alignment = Alignment(horizontal='left', vertical='center')

        # Row 22: When to sell
        start_sheet['A22'] = "When to sell? When a company has exceeded its Intrinsic Value, it is in sell territory. If you want to hold, look at the 10-year trend on EPS. Stretch out those earnings over the next 10 years and compare them to the gain one would receive if you sold a share of the stock at the current price and invested it at the AAA Corporate bond rate for the next ten years. If you make more with the bond, it's a good time to sell the stock and look for another investment."
        start_sheet['A22'].alignment = Alignment(horizontal='left', vertical='center')

        # Row 24: AAA Corporate Bond
        start_sheet.merge_cells('A24:B24')
        start_sheet['A24'] = "AAA Corporate Bond"
        start_sheet['A24'].font = Font(bold=True)
        start_sheet['A24'].alignment = Alignment(horizontal='center', vertical='center')
        aaa_yield = get_aaa_yield(FRED_API_KEY)
        start_sheet['C24'] = aaa_yield
        start_sheet['C24'].number_format = '0.00%'
        start_sheet['C24'].alignment = Alignment(horizontal='center', vertical='center')

        # Row 26: Graham Criteria for Investment Grade Stocks
        start_sheet.merge_cells('A26:D26')
        start_sheet['A26'] = "Graham Criteria for Investment Grade Stocks"
        start_sheet['A26'].font = Font(bold=True)
        start_sheet['A26'].alignment = Alignment(horizontal='center', vertical='center')

        # Rows 28 to 34: Investment Grade Criteria
        investment_criteria = [
            "1) Adequate Size = 500 million in annual sales and 100 million in assets.",
            "2) Strong Financially = 2:1 Current Yield.",
            "3) 20 Years of Dividend Payment (I use 10 years for simplicity).",
            "4) No negative Earnings Per Share over the last 10 years.",
            "5) Earnings growth of at least 33% over last 10 years. (I use a CAGR of at least 3%).",
            "6) Share Price <= 1.5x Net Asset Value.",
            "7) Share Price <= 15x Average Earnings for past 3 years."
        ]
        for i, crit in enumerate(investment_criteria, start=28):
            start_sheet[f'A{i}'] = crit
            start_sheet[f'A{i}'].alignment = Alignment(horizontal='left', vertical='center')

        # Row 36: Graham Criteria for Good Value Companies
        start_sheet.merge_cells('A36:E36')
        start_sheet['A36'] = "Graham criteria for good value companies. 7/10 is goal."
        start_sheet['A36'].font = Font(bold=True)
        start_sheet['A36'].alignment = Alignment(horizontal='center', vertical='center')

        # Rows 38 to 47: Value Company Criteria
        value_criteria = [
            "1) Earnings/Price ratio = 2x AAA Corporate Bond Yield.",
            "2) Price/Earnings ratio = .4 of highest P/E average of the last 10 years.",
            "3) Dividend Yield >= 2/3 of AAA Bond Yield.",
            "4) Share Price = 2/3 Tangible Book Value per Share.",
            "5) 2/3 Net Current Asset Value.",
            "6) Total Debt Lower than Tangible Book Value.",
            "7) Current Ratio >= 2.",
            "8) Total Debt <= Net Quick Liquidation Value.",
            "9) Earnings have doubled in the last 10 years.",
            "10) Earnings have declined no more than 5% in the last 10 years."
        ]
        for i, crit in enumerate(value_criteria, start=38):
            start_sheet[f'A{i}'] = crit
            start_sheet[f'A{i}'].alignment = Alignment(horizontal='left', vertical='center')

        # Row 48: Margin of Safety
        start_sheet.merge_cells('A48:B48')
        start_sheet['A48'] = "Margin of Safety"
        start_sheet['A48'].font = Font(bold=True)
        start_sheet['A48'].alignment = Alignment(horizontal='center', vertical='center')
        margin_of_safety = self.margin_of_safety_var.get()
        start_sheet['C48'] = margin_of_safety / 100  # Convert to decimal for percentage format
        start_sheet['C48'].number_format = '0.00%'
        start_sheet['C48'].alignment = Alignment(horizontal='center', vertical='center')

        # Disable text wrapping for all rows
        for row in start_sheet.iter_rows():
            for cell in row:
                if cell.row in [24, 26, 36, 48]:  # Updated to include row 48
                    cell.alignment = Alignment(wrap_text=False, horizontal='center', vertical='center')
                else:
                    cell.alignment = Alignment(wrap_text=False, horizontal='left', vertical='center')

    def get_tangible_book_value_per_share(key_metrics_data: list) -> float:
        """Extract latest tangible book value per share from key_metrics_data."""
        if not key_metrics_data:
            return 0.0
        latest = key_metrics_data[0]
        tbvps = latest.get('tangibleBookValuePerShare', 0.0)
        try:
            return float(tbvps)
        except (ValueError, TypeError):
            return 0.0

    def export_qualifying_stocks(self, exchange, min_criteria):
        analyze_logger.debug(f"Exporting {exchange} qualifying stocks with min_criteria={min_criteria}")

        conn, cursor = get_stocks_connection()
        try:
            cursor.execute("""
                SELECT s.ticker, s.company_name, g.common_score, s.date, s.roe, s.rotc, s.eps, s.dividend,
                    s.ticker_list_hash, s.balance_data, s.timestamp, s.debt_to_equity, s.eps_ttm,
                    s.book_value_per_share, s.latest_revenue, s.available_data_years, s.sector,
                    s.years, s.latest_total_assets, s.latest_total_liabilities, s.latest_shares_outstanding,
                    s.latest_long_term_debt, s.latest_short_term_debt, s.latest_current_assets,
                    s.latest_current_liabilities, s.latest_book_value, s.historic_pe_ratios,
                    s.latest_net_income, s.eps_cagr, s.latest_free_cash_flow, s.raw_income_data,
                    s.raw_balance_data, s.raw_dividend_data, s.raw_profile_data, s.raw_cash_flow_data,
                    s.raw_key_metrics_data
                FROM stocks s
                JOIN graham_qualifiers g ON s.ticker = g.ticker
                WHERE g.exchange = ? AND g.common_score >= ?
                ORDER BY g.common_score DESC, s.ticker ASC
            """, (exchange, min_criteria))
            results = cursor.fetchall()
            if not results:
                messagebox.showinfo("No Results", f"No {exchange} stocks meet the Graham criteria with minimum score {min_criteria}.")
                analyze_logger.info(f"No qualifying {exchange} stocks found with min_criteria={min_criteria}")
                return

            tickers = [row[0] for row in results]
            prices = self.fetch_current_prices(tickers)

            # Create workbook
            wb = openpyxl.Workbook()

            # "Start Here" sheet
            start_sheet = wb.active
            start_sheet.title = "Start Here"
            self.setup_start_here_sheet(start_sheet)

            # Process stock data for summary tabs
            stock_data_list = []
            for row in results:
                ticker, company_name, common_score, date, roe, rotc, eps, dividend, ticker_list_hash, balance_data, timestamp, debt_to_equity, eps_ttm, book_value_per_share, latest_revenue, available_data_years, sector, years, latest_total_assets, latest_total_liabilities, latest_shares_outstanding, latest_long_term_debt, latest_short_term_debt, latest_current_assets, latest_current_liabilities, latest_book_value, historic_pe_ratios, latest_net_income, eps_cagr, latest_free_cash_flow, raw_income_data, raw_balance_data, raw_dividend_data, raw_profile_data, raw_cash_flow_data, raw_key_metrics_data = row
                price = prices.get(ticker, "N/A")
                if price == "N/A" or not isinstance(price, (int, float)):
                    continue  # Skip stocks without valid price data
                try:
                    eps_list = [float(x) for x in eps.split(",")] if eps else []
                    div_list = [float(x) for x in dividend.split(",")] if dividend else []
                    balance_data_dict = json.loads(balance_data) if balance_data else []
                except Exception as e:
                    analyze_logger.error(f"Error parsing data for {ticker}: {e}")
                    continue

                # Compute tangible BVPS from raw_key_metrics_data
                tangible_bvps = get_tangible_book_value_per_share(json.loads(raw_key_metrics_data) if raw_key_metrics_data else [])

                intrinsic_value = self.calculate_intrinsic_value({'eps_ttm': eps_ttm, 'eps_cagr': eps_cagr})
                if pd.isna(intrinsic_value) or intrinsic_value == 0:
                    margin_of_safety = "N/A"
                else:
                    margin_of_safety = (intrinsic_value - price) / intrinsic_value * 100  # Percent current price is below/above intrinsic value
                expected_return = self.expected_return_var.get() / 100
                buy_price = intrinsic_value * (1 - (self.margin_of_safety_var.get() / 100)) if not pd.isna(intrinsic_value) else "N/A"
                graham_score = calculate_graham_score_8(ticker, price, None, None, debt_to_equity, eps_list, div_list, {}, balance_data_dict, json.loads(raw_key_metrics_data) if raw_key_metrics_data else [], available_data_years, latest_revenue, sector)

                stock_data = {
                    "company_name": company_name,
                    "ticker": ticker,
                    "sector": sector,
                    "mos": margin_of_safety if margin_of_safety != "N/A" else "N/A",
                    "graham_score": graham_score if graham_score is not None else "N/A",
                    "current_price": price,
                    "intrinsic_value": intrinsic_value if not pd.isna(intrinsic_value) else "N/A",
                    "buy_price": buy_price if buy_price != "N/A" else "N/A",
                    "latest_total_assets": latest_total_assets,
                    "latest_total_liabilities": latest_total_liabilities,
                    "latest_shares_outstanding": latest_shares_outstanding,
                    "latest_current_assets": latest_current_assets,
                    "latest_current_liabilities": latest_current_liabilities,
                    "free_cash_flow": latest_free_cash_flow,
                    "raw_key_metrics_data": raw_key_metrics_data,  # Added for bank_metrics
                    "tangible_book_value_per_share": tangible_bvps  # Added for P/TBV
                }
                stock_data_list.append(stock_data)

            # Sort stock data list by company name
            stock_data_list.sort(key=lambda x: x["company_name"])

            # Define financial sectors
            financial_sectors = ['Financial Services', 'Finance', 'Banking', 'financials']

            # Make case-insensitive comparison
            financial_sectors_lower = [s.lower() for s in financial_sectors]
            financial_stocks = [stock for stock in stock_data_list if stock['sector'].lower() in financial_sectors_lower]
            other_stocks = [stock for stock in stock_data_list if stock['sector'].lower() not in financial_sectors_lower]

            # Headers for summary sheets
            headers = [
                "Company Name", "Ticker", "Sector", "Bargain?", "MOS", "Graham Score", "Stability Test",
                "ROE>12%", "ROTC>12%", "EPS Uptrend", "LTD<5 Years", "Dividend", "Buyback",
                "POT #1", "POT #2", "Current Price", "Intrinsic Value", "Buy Price"
            ]

            # Function to create a summary sheet
            def create_summary_sheet(sheet_name, stocks, is_financial=False):
                sheet = wb.create_sheet(sheet_name)
                headers = ["Company Name", "Ticker", "Sector", "Bargain?", "MOS", "Graham Score", "Stability Test", "ROE>12%", "ROTC>12%", "EPS Uptrend", "LTD<5 Years", "Dividend", "Buyback", "POT #1", "POT #2", "Current Price", "Intrinsic Value", "Buy Price"]
                if is_financial:
                    headers.extend(["ROA", "ROE", "NIM", "P/TBV"])  # Extra columns for financials
                
                for col, header in enumerate(headers, start=1):
                    cell = sheet.cell(row=1, column=col, value=header)
                    cell.font = Font(size=12, bold=True)
                    cell.alignment = Alignment(horizontal='center', vertical='center')

                # Freeze the header row (row 1) and Ticker column (column A)
                sheet.freeze_panes = 'B2'

                # Populate rows
                for row_idx, stock in enumerate(stocks, start=2):
                    company_cell = sheet.cell(row=row_idx, column=1, value=stock["company_name"])
                    company_cell.hyperlink = f"#'{stock['ticker']}'!A1"
                    company_cell.style = "Hyperlink"
                    sheet.cell(row=row_idx, column=2, value=stock["ticker"])
                    sheet.cell(row=row_idx, column=3, value=stock["sector"])
                    # "Bargain?" column added after "Sector"
                    sheet.cell(row=row_idx, column=4).value = f"='{stock['ticker']}'!M26"
                    sheet.cell(row=row_idx, column=5, value=stock["mos"] / 100 if stock["mos"] != "N/A" else "N/A")
                    sheet.cell(row=row_idx, column=6, value=stock["graham_score"] if stock["graham_score"] != "N/A" else "N/A")
                    sheet.cell(row=row_idx, column=7).value = f"='{stock['ticker']}'!G45"  # Stability Test
                    sheet.cell(row=row_idx, column=8).value = f"='{stock['ticker']}'!L5"   # ROE>12%
                    sheet.cell(row=row_idx, column=9).value = f"='{stock['ticker']}'!L6"   # ROTC>12%
                    sheet.cell(row=row_idx, column=10).value = (
                        f"=IF(SUMPRODUCT(--('{stock['ticker']}'!C7:'{stock['ticker']}'!K7 < '{stock['ticker']}'!B7:'{stock['ticker']}'!J7))<=2, "
                        f"\"Yes\", IF(SUMPRODUCT(--('{stock['ticker']}'!C7:'{stock['ticker']}'!K7 < '{stock['ticker']}'!B7:'{stock['ticker']}'!J7))=3, "
                        f"\"Maybe\", \"No\"))"
                    )  # EPS Uptrend
                    sheet.cell(row=row_idx, column=11).value = f"='{stock['ticker']}'!G13"  # LTD<5 Years
                    sheet.cell(row=row_idx, column=12).value = f"='{stock['ticker']}'!K10"  # Dividend
                    sheet.cell(row=row_idx, column=13).value = f"='{stock['ticker']}'!L29"  # Buyback
                    sheet.cell(row=row_idx, column=14).value = f"='{stock['ticker']}'!C22"  # POT #1
                    sheet.cell(row=row_idx, column=15).value = f"='{stock['ticker']}'!C27"  # POT #2
                    sheet.cell(row=row_idx, column=16, value=stock["current_price"] if stock["current_price"] != "N/A" else "N/A")
                    sheet.cell(row=row_idx, column=17, value=stock["intrinsic_value"] if stock["intrinsic_value"] != "N/A" else "N/A")
                    sheet.cell(row=row_idx, column=18, value=stock["buy_price"] if stock["buy_price"] != "N/A" else "N/A")
                    
                    if is_financial:
                        # Compute financial metrics
                        bank_metrics = get_bank_metrics(json.loads(stock["raw_key_metrics_data"]) if "raw_key_metrics_data" in stock else [])
                        tangible_bvps = stock.get("tangible_book_value_per_share", 0)
                        price = stock.get("current_price", 1)
                        ptbv = price / tangible_bvps if tangible_bvps > 0 else "N/A"
                        sheet.cell(row=row_idx, column=19, value=bank_metrics.get('roa', "N/A"))
                        sheet.cell(row=row_idx, column=20, value=bank_metrics.get('roe', "N/A"))
                        sheet.cell(row=row_idx, column=21, value=bank_metrics.get('netInterestMargin', "N/A"))
                        sheet.cell(row=row_idx, column=22, value=ptbv)

                # Set number formats for numeric columns
                last_row = len(stocks) + 1
                for row in range(2, last_row + 1):
                    sheet.cell(row=row, column=5).number_format = '0.00%'  # MOS
                    sheet.cell(row=row, column=6).number_format = '0'      # Graham Score
                    sheet.cell(row=row, column=8).number_format = '0.00%'  # ROE>12%
                    sheet.cell(row=row, column=9).number_format = '0.00%'  # ROTC>12%
                    sheet.cell(row=row, column=11).number_format = '0.00'  # LTD<5 Years
                    sheet.cell(row=row, column=12).number_format = '$#,##0.00'  # Dividend
                    sheet.cell(row=row, column=16).number_format = '$#,##0.00'  # Current Price
                    sheet.cell(row=row, column=17).number_format = '$#,##0.00'  # Intrinsic Value
                    sheet.cell(row=row, column=18).number_format = '$#,##0.00'  # Buy Price
                    if is_financial:
                        sheet.cell(row=row, column=19).number_format = '0.00%'  # ROA
                        sheet.cell(row=row, column=20).number_format = '0.00%'  # ROE
                        sheet.cell(row=row, column=21).number_format = '0.00%'  # NIM
                        sheet.cell(row=row, column=22).number_format = '0.00'   # P/TBV

                # Define fill colors
                green_fill = PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')
                light_blue_fill = PatternFill(start_color='ADD8E6', end_color='ADD8E6', fill_type='solid')

                # Apply conditional formatting for "Bargain?" column (column D)
                sheet.conditional_formatting.add(
                    f'D2:D{last_row}',
                    CellIsRule(operator='equal', formula=['"Yes"'], fill=green_fill)
                )
                sheet.conditional_formatting.add(
                    f'D2:D{last_row}',
                    CellIsRule(operator='equal', formula=['"Maybe"'], fill=light_blue_fill)
                )

                # Stability Test (now column G)
                sheet.conditional_formatting.add(
                    f'G2:G{last_row}',
                    FormulaRule(formula=['G2>=8'], fill=green_fill)
                )
                sheet.conditional_formatting.add(
                    f'G2:G{last_row}',
                    FormulaRule(formula=['AND(G2>=6, G2<8)'], fill=light_blue_fill)
                )

                # ROE>12% (now column H)
                sheet.conditional_formatting.add(
                    f'H2:H{last_row}',
                    FormulaRule(formula=['H2>=0.12'], fill=green_fill)
                )
                sheet.conditional_formatting.add(
                    f'H2:H{last_row}',
                    FormulaRule(formula=['AND(H2>=0.10, H2<0.12)'], fill=light_blue_fill)
                )

                # ROTC>12% (now column I)
                sheet.conditional_formatting.add(
                    f'I2:I{last_row}',
                    FormulaRule(formula=['I2>=0.12'], fill=green_fill)
                )
                sheet.conditional_formatting.add(
                    f'I2:I{last_row}',
                    FormulaRule(formula=['AND(I2>=0.10, I2<0.12)'], fill=light_blue_fill)
                )

                # EPS Uptrend (now column J)
                sheet.conditional_formatting.add(
                    f'J2:J{last_row}',
                    CellIsRule(operator='equal', formula=['"Yes"'], fill=green_fill)
                )
                sheet.conditional_formatting.add(
                    f'J2:J{last_row}',
                    CellIsRule(operator='equal', formula=['"Maybe"'], fill=light_blue_fill)
                )

                # LTD<5 Years (now column K)
                sheet.conditional_formatting.add(
                    f'K2:K{last_row}',
                    FormulaRule(formula=['K2<=5'], fill=green_fill)
                )

                # Buyback (now column M)
                sheet.conditional_formatting.add(
                    f'M2:M{last_row}',
                    CellIsRule(operator='equal', formula=['"Yes"'], fill=green_fill)
                )

                # POT #1 (now column N)
                sheet.conditional_formatting.add(
                    f'N2:N{last_row}',
                    CellIsRule(operator='equal', formula=['"Yes"'], fill=green_fill)
                )

                # POT #2 (now column O)
                sheet.conditional_formatting.add(
                    f'O2:O{last_row}',
                    CellIsRule(operator='equal', formula=['"Yes"'], fill=green_fill)
                )

                # Set column widths
                column_widths = [55, 12, 25, 14, 10, 19, 18, 14, 16, 18, 17, 14, 14, 12, 12, 18, 20, 15]
                if is_financial:
                    column_widths.extend([10, 10, 10, 10])  # Extra widths for ROA, ROE, NIM, P/TBV
                for col, width in enumerate(column_widths, start=1):
                    sheet.column_dimensions[get_column_letter(col)].width = width

                # Set alignment
                for row in sheet.iter_rows():
                    for cell in row:
                        cell.alignment = Alignment(horizontal='center', vertical='center')

                # Auto-filter range includes all columns (A to R or more for financial)
                last_col_letter = get_column_letter(len(headers))
                sheet.auto_filter.ref = f"A1:{last_col_letter}{last_row}"
                sheet.auto_filter.add_sort_condition(f"A2:A{last_row}")

            # Update the calls to create_summary_sheet
            # Create 'Winning Stocks' sheet for non-financial stocks
            if other_stocks:
                create_summary_sheet("Winning Stocks", other_stocks)
                print(f"Created 'Winning Stocks' sheet with {len(other_stocks)} stocks")

            # Create 'Financial Winners' sheet for financial stocks
            if financial_stocks:
                create_summary_sheet("Financial Winners", financial_stocks, is_financial=True)
                print(f"Created 'Financial Winners' sheet with {len(financial_stocks)} stocks")
            else:
                print("No financial stocks to create 'Financial Winners' sheet")

            # Sort results by ticker for individual sheets
            results.sort(key=lambda x: x[0])  # Assuming ticker is the first element

            # Calculate factor for M27 formula based on margin of safety
            margin_of_safety = self.margin_of_safety_var.get()
            factor = 1 - (margin_of_safety / 100)

            # Create individual stock sheets
            for row in results:
                ticker, company_name, common_score, date, roe, rotc, eps, dividend, ticker_list_hash, balance_data, timestamp, debt_to_equity, eps_ttm, book_value_per_share, latest_revenue, available_data_years, sector, years, latest_total_assets, latest_total_liabilities, latest_shares_outstanding, latest_long_term_debt, latest_short_term_debt, latest_current_assets, latest_current_liabilities, latest_book_value, historic_pe_ratios, latest_net_income, eps_cagr, latest_free_cash_flow, raw_income_data, raw_balance_data, raw_dividend_data, raw_profile_data, raw_cash_flow_data, raw_key_metrics_data = row
                price = prices.get(ticker, "N/A")
                stock_sheet = wb.create_sheet(ticker[:31])  # Truncate to 31 chars for Excel compatibility

                # Set row heights to 15 for rows 1 to 55
                for row_num in range(1, 56):
                    stock_sheet.row_dimensions[row_num].height = 15

                # Merge cells A2:K3 for company name and ticker
                stock_sheet.merge_cells('A2:K3')
                company_ticker_cell = stock_sheet['A2']
                company_ticker_cell.value = f"{company_name.upper()} ({ticker})"
                company_ticker_cell.font = Font(bold=True, size=18)
                company_ticker_cell.alignment = Alignment(horizontal='center', vertical='center')
                company_ticker_cell.hyperlink = f"https://www.morningstar.com/stocks/xnys/{ticker}/quote"

                # Set column widths
                stock_sheet.column_dimensions['A'].width = 30
                for col in range(2, 13):  # B to L
                    stock_sheet.column_dimensions[get_column_letter(col)].width = 10
                stock_sheet.column_dimensions['M'].width = 15
                stock_sheet.column_dimensions['N'].width = 10
                stock_sheet.column_dimensions['O'].width = 15

                # Determine last updated date and format as "31-May-25"
                ticker_file = NYSE_LIST_FILE if exchange == "NYSE" else NASDAQ_LIST_FILE
                if os.path.exists(ticker_file):
                    last_updated = datetime.fromtimestamp(os.path.getmtime(ticker_file)).strftime('%d-%b-%y')
                else:
                    last_updated = "Unknown"

                # Merge L1:N1 for "Last Updated:"
                stock_sheet.merge_cells('L1:N1')
                last_updated_label = stock_sheet['L1']
                last_updated_label.value = "Last Updated:"
                last_updated_label.alignment = Alignment(horizontal='center', vertical='center')

                # Set O1 to last updated date
                stock_sheet['O1'].value = last_updated
                stock_sheet['O1'].alignment = Alignment(horizontal='center', vertical='center')

                # Fetch AAA yield
                aaa_yield = get_aaa_yield(FRED_API_KEY)

                # Calculate P/E ratio
                if isinstance(price, (int, float)) and eps_ttm and eps_ttm > 0:
                    pe_ratio = price / eps_ttm
                else:
                    pe_ratio = "N/A"

                # Set L2: current price
                cell_l2 = stock_sheet['L2']
                if isinstance(price, (int, float)):
                    cell_l2.value = price
                    cell_l2.number_format = '$#,##0.00'
                else:
                    cell_l2.value = "N/A"
                cell_l2.font = Font(bold=True)
                cell_l2.alignment = Alignment(horizontal='center', vertical='center')

                # Set M2: "Current Price"
                cell_m2 = stock_sheet['M2']
                cell_m2.value = "Current Price"
                cell_m2.font = Font(bold=True)
                cell_m2.alignment = Alignment(horizontal='center', vertical='center')

                # Set N2: P/E ratio
                cell_n2 = stock_sheet['N2']
                if isinstance(pe_ratio, (int, float)):
                    cell_n2.value = pe_ratio
                    cell_n2.number_format = '0.00'
                else:
                    cell_n2.value = "N/A"
                cell_n2.font = Font(bold=True)
                cell_n2.alignment = Alignment(horizontal='center', vertical='center')

                # Set O2: "P/E Ratio"
                cell_o2 = stock_sheet['O2']
                cell_o2.value = "P/E Ratio"
                cell_o2.font = Font(bold=True)
                cell_o2.alignment = Alignment(horizontal='center', vertical='center')

                # Set L3: AAA yield
                cell_l3 = stock_sheet['L3']
                if isinstance(aaa_yield, (int, float)):
                    cell_l3.value = aaa_yield
                    cell_l3.number_format = '0.00%'
                else:
                    cell_l3.value = "N/A"
                cell_l3.font = Font(bold=True)
                cell_l3.alignment = Alignment(horizontal='center', vertical='center')

                # Merge M3:O3 and set "AAA Corporate Bond Rate"
                stock_sheet.merge_cells('M3:O3')
                cell_m3 = stock_sheet['M3']
                cell_m3.value = "AAA Corporate Bond Rate"
                cell_m3.font = Font(bold=True)
                cell_m3.alignment = Alignment(horizontal='center', vertical='center')

                # Set thick outside borders for L2:O3
                bold_side = Side(style='thick')
                thin_side = Side(style='thin')
                cell_l2.border = Border(top=bold_side, left=bold_side, bottom=thin_side)
                cell_m2.border = Border(top=bold_side, bottom=thin_side, right=bold_side, left=thin_side)
                cell_n2.border = Border(top=bold_side, bottom=thin_side, left=bold_side)
                cell_o2.border = Border(top=bold_side, bottom=thin_side, right=bold_side, left=thin_side)
                cell_l3.border = Border(top=thin_side, left=bold_side, bottom=bold_side, right=thin_side)
                cell_m3.border = Border(top=thin_side, left=thin_side, bottom=bold_side)

                # Set thick border for N4:O4
                for col in ['N', 'O']:
                    cell = stock_sheet[f'{col}4']
                    cell.border = Border(top=bold_side)

                # Set thin bottom border for P3
                for col in ['P']:
                    cell = stock_sheet[f'{col}3']
                    cell.border = Border(left=bold_side)

                # Define labels with Unicode subscripts
                sub_1 = '\u2081'
                sub_0 = '\u2080'
                sub_10 = sub_1 + sub_0
                labels = [
                    "Year",
                    f"ROE{sub_10}",
                    f"ROTC{sub_10}",
                    f"EPS{sub_10}",
                    f"EPS{sub_10} CAGR",  # Updated to CAGR
                    f"EPS{sub_10} Proj",
                    f"DIV{sub_10}",
                    f"DIV{sub_10} CAGR",  # Updated to CAGR
                    f"DIV{sub_10} Proj"
                ]

                # Set labels in A4:A12 and M4:M12
                for i, label in enumerate(labels, start=4):
                    cell_a = stock_sheet[f'A{i}']
                    cell_m = stock_sheet[f'M{i}']
                    cell_a.value = label
                    cell_m.value = label
                    cell_a.font = Font(size=10, bold=True)
                    cell_m.font = Font(size=10, bold=True)
                    cell_a.alignment = Alignment(horizontal='center', vertical='center')
                    cell_m.alignment = Alignment(horizontal='center', vertical='center')

                # Parse historical data
                years_list = [int(y) for y in years.split(",")] if years else []
                roe_list = [float(x) for x in roe.split(",")] if roe else []
                rotc_list = [float(x) for x in rotc.split(",")] if rotc else []
                eps_list = [float(x) for x in eps.split(",")] if eps else []
                div_list = [float(x) for x in dividend.split(",")] if dividend else []

                # Take the last 10 years
                years_list = years_list[-10:]
                roe_list = roe_list[-10:]
                rotc_list = rotc_list[-10:]
                eps_list = eps_list[-10:]
                div_list = div_list[-10:]

                # Populate B4:K4 with years (bold)
                for col, year in enumerate(years_list, start=2):  # B to K
                    cell = stock_sheet.cell(row=4, column=col)
                    cell.value = year
                    cell.font = Font(bold=True)

                # Populate B5:K5 with ROE (as decimal)
                for col, roe_value in enumerate(roe_list, start=2):
                    cell = stock_sheet.cell(row=5, column=col)
                    if col-2 < len(roe_list):
                        cell.value = roe_value / 100 if isinstance(roe_value, (int, float)) else "N/A"
                    else:
                        cell.value = "N/A"
                    if isinstance(cell.value, (int, float)):
                        cell.number_format = '0.00%'

                # Populate B6:K6 with ROTC (as decimal)
                for col, rotc_value in enumerate(rotc_list, start=2):
                    cell = stock_sheet.cell(row=6, column=col)
                    if col-2 < len(rotc_list):
                        cell.value = rotc_value / 100 if isinstance(rotc_value, (int, float)) else "N/A"
                    else:
                        cell.value = "N/A"
                    if isinstance(cell.value, (int, float)):
                        cell.number_format = '0.00%'

                # Populate B7:K7 with EPS
                for col, eps_value in enumerate(eps_list, start=2):
                    cell = stock_sheet.cell(row=7, column=col)
                    cell.value = eps_value if col-2 < len(eps_list) else "N/A"
                    if isinstance(cell.value, (int, float)):
                        cell.number_format = '$#,##0.00'

                # Populate B10:K10 with Dividends
                for col, div_value in enumerate(div_list, start=2):
                    cell = stock_sheet.cell(row=10, column=col)
                    cell.value = div_value if col-2 < len(div_list) else "N/A"
                    if isinstance(cell.value, (int, float)):
                        cell.number_format = '$#,##0.00'

                # Set L4 to "Avg10"
                stock_sheet['L4'].value = f"Avg{sub_10}"
                stock_sheet['L4'].font = Font(bold=True)
                stock_sheet['L4'].alignment = Alignment(horizontal='center', vertical='center')

                # Set averages using Excel formulas for L5:L7
                stock_sheet['L5'].value = "=AVERAGE(B5:K5)"
                stock_sheet['L5'].number_format = '0.00%'
                stock_sheet['L6'].value = "=AVERAGE(B6:K6)"
                stock_sheet['L6'].number_format = '0.00%'
                stock_sheet['L7'].value = "=AVERAGE(B7:K7)"
                stock_sheet['L7'].number_format = '$#,##0.00'

                # Set L8 to EPS CAGR
                stock_sheet['L8'].value = (
                    '=IF(AND(COUNT(B7:K7)>=2, B7<>0), '
                    '(INDEX(B7:K7, COUNT(B7:K7)) / B7)^(1/(COUNT(B7:K7)-1)) - 1, "N/A")'
                )
                stock_sheet['L8'].number_format = '0.00%'

                #Set averages using Excel formulas for L9:10
                stock_sheet['L9'].value = "=AVERAGE(B9:K9)"
                stock_sheet['L9'].number_format = '$#,##0.00'
                stock_sheet['L10'].value = "=AVERAGE(B10:K10)"
                stock_sheet['L10'].number_format = '$#,##0.00'

                # Set L11 to DIV CAGR
                stock_sheet['L11'].value = (
                    '=IF(AND(COUNT(B10:K10)>=2, B10<>0), '
                    '(INDEX(B10:K10, COUNT(B10:K10)) / B10)^(1/(COUNT(B10:K10)-1)) - 1, "N/A")'
                )
                stock_sheet['L11'].number_format = '0.00%'

                #Set average using Excel formula for L12
                stock_sheet['L12'].value = "=AVERAGE(B12:K12)"
                stock_sheet['L12'].number_format = '$#,##0.00'

                # Set EPS projections
                stock_sheet['B9'].value = f"={get_column_letter(11)}7 * (1 + {get_column_letter(12)}8)"
                stock_sheet['B9'].number_format = '$#,##0.00'
                for col in range(3, 12):  # C to K
                    prev_col_letter = get_column_letter(col - 1)
                    stock_sheet.cell(row=9, column=col).value = f"={prev_col_letter}9 * (1 + ${get_column_letter(12)}$8)"
                    stock_sheet.cell(row=9, column=col).number_format = '$#,##0.00'

                # Set Dividend projections
                stock_sheet['B12'].value = f"={get_column_letter(11)}10 * (1 + {get_column_letter(12)}11)"
                stock_sheet['B12'].number_format = '$#,##0.00'
                for col in range(3, 12):  # C to K
                    prev_col_letter = get_column_letter(col - 1)
                    stock_sheet.cell(row=12, column=col).value = f"={prev_col_letter}12 * (1 + ${get_column_letter(12)}$11)"
                    stock_sheet.cell(row=12, column=col).number_format = '$#,##0.00'

                # Row 13
                stock_sheet['A13'].value = "Long Term Debt ($M)"
                stock_sheet['A13'].font = Font(bold=True)
                stock_sheet['A13'].alignment = Alignment(horizontal='center', vertical='center')

                if latest_long_term_debt is not None:
                    stock_sheet['B13'].value = latest_long_term_debt / 1_000_000
                    stock_sheet['B13'].number_format = '$#,##0'
                else:
                    stock_sheet['B13'].value = "N/A"

                stock_sheet.merge_cells('C13:D13')
                stock_sheet['C13'].value = "Net Income ($M)"
                stock_sheet['C13'].font = Font(bold=True)
                stock_sheet['C13'].alignment = Alignment(horizontal='center', vertical='center')

                if latest_net_income is not None:
                    stock_sheet['E13'].value = latest_net_income / 1_000_000
                    stock_sheet['E13'].number_format = '$#,##0'
                else:
                    stock_sheet['E13'].value = "N/A"

                stock_sheet['F13'].value = "LTD/NI"
                stock_sheet['F13'].font = Font(bold=True)
                stock_sheet['F13'].alignment = Alignment(horizontal='center', vertical='center')

                stock_sheet['G13'].value = "=IF(E13=0, \"N/A\", B13/E13)"
                stock_sheet['G13'].number_format = '0.00'

                # Set alignment for row 13
                for col in ['A', 'B', 'C', 'E', 'F', 'G']:
                    stock_sheet[f'{col}13'].alignment = Alignment(horizontal='center', vertical='center')

                # Set thick bottom border for B14:C14
                for col in ['B', 'C']:
                    cell = stock_sheet[f'{col}14']
                    cell.border = Border(bottom=bold_side)

                # Set thick side border for D15
                for col in ['D']:
                    cell = stock_sheet[f'{col}15']
                    cell.border = Border(left=bold_side)

                # Set H14
                stock_sheet['H14'].value = "Intelligent Investor Earnings Multiplier (8.5-20)"
                stock_sheet['H14'].font = Font(bold=True)
                stock_sheet['H14'].alignment = Alignment(horizontal='left', vertical='center')

                # Calculate earnings multiplier
                if eps_cagr is not None:
                    g = max(eps_cagr * 100, 0)
                    earnings_multiplier = min(8.5 + 2 * g, 20)
                else:
                    earnings_multiplier = "N/A"

                # Set M14
                if earnings_multiplier != "N/A":
                    stock_sheet['M14'].value = earnings_multiplier
                    stock_sheet['M14'].number_format = '0.00'
                else:
                    stock_sheet['M14'].value = "N/A"
                stock_sheet['M14'].alignment = Alignment(horizontal='center', vertical='center')

                # Set thin bottom border for H14:M14
                for col in ['H','I', 'J', 'K', 'L', 'M']:
                    cell = stock_sheet[f'{col}14']
                    cell.border = Border(bottom=thin_side)

                # Merge A15:C15 and set thick outside borders
                stock_sheet.merge_cells('A15:C15')
                cell_a15 = stock_sheet['A15']
                cell_a15.value = "Value to a Private Owner Test (POT) (Page 98)"
                cell_a15.font = Font(bold=True)
                cell_a15.alignment = Alignment(horizontal='center', vertical='center')
                cell_a15.border = Border(left=bold_side, top=bold_side, right=bold_side, bottom=bold_side)

                # Set H15
                stock_sheet['H15'].value = "Intrinsic Value"
                stock_sheet['H15'].font = Font(bold=True)
                stock_sheet['H15'].alignment = Alignment(horizontal='left', vertical='center')

                # Set M15: Intrinsic Value based on latest EPS, multiplier, normalization factor, and AAA yield
                if isinstance(stock_sheet['M14'].value, (int, float)) and isinstance(stock_sheet['L3'].value, (int, float)):
                    stock_sheet['M15'].value = "=(K7 * M14 * 4.4) / (100 * L3)"
                    stock_sheet['M15'].number_format = '$#,##0.00'
                else:
                    stock_sheet['M15'].value = "N/A"
                stock_sheet['M15'].alignment = Alignment(horizontal='center', vertical='center')

                # Set double thin bottom border for H15:M15
                double_side = Side(style='double')
                for col in ['H','I', 'J', 'K', 'L', 'M']:
                    cell = stock_sheet[f'{col}15']
                    cell.border = Border(bottom=double_side)

                # Set A16:A27 labels, left-aligned and vertically centered
                a_labels = [
                    "Total Assets ($M)",
                    "Total Liabilities ($M)",
                    "Free Cash Flow ($M)",
                    "Shares Outstanding (M)",
                    "Value Per Share of Cash and Assets:",
                    "Price in Relation to POT >100%",
                    "Is the Company Worth Owning?",
                    "",  # Skipped for spacing
                    "Working Capital ($M)",
                    "Value Per Share of Working Capital:",
                    "Price in Relation to Working Capital >100%:",
                    "Does Capital Alone Make It A Bargain?"
                ]
                for i, label in enumerate(a_labels, start=16):
                    cell = stock_sheet[f'A{i}']
                    cell.value = label
                    cell.alignment = Alignment(horizontal='left', vertical='center')

                # Populate C16:C19 and C24 with data, centered horizontally and vertically
                if latest_total_assets is not None:
                    stock_sheet['C16'].value = latest_total_assets / 1_000_000
                    stock_sheet['C16'].number_format = '$#,##0'
                else:
                    stock_sheet['C16'].value = "N/A"
                stock_sheet['C16'].alignment = Alignment(horizontal='center', vertical='center')

                if latest_total_liabilities is not None:
                    stock_sheet['C17'].value = latest_total_liabilities / 1_000_000
                    stock_sheet['C17'].number_format = '$#,##0'
                else:
                    stock_sheet['C17'].value = "N/A"
                stock_sheet['C17'].alignment = Alignment(horizontal='center', vertical='center')

                # C18: Free Cash Flow ($M)
                analyze_logger.debug(f"{ticker}: Raw latest_free_cash_flow from DB = {latest_free_cash_flow}")
                if isinstance(latest_free_cash_flow, (int, float)) and latest_free_cash_flow is not None:
                    stock_sheet['C18'].value = latest_free_cash_flow / 1_000_000
                    stock_sheet['C18'].number_format = '$#,##0'
                    analyze_logger.debug(f"{ticker}: Set FCF in C18 = {latest_free_cash_flow / 1_000_000} $M")
                else:
                    stock_sheet['C18'].value = "N/A"
                    analyze_logger.warning(f"{ticker}: FCF unavailable or invalid for C18 (value: {latest_free_cash_flow})")
                stock_sheet['C18'].alignment = Alignment(horizontal='center', vertical='center')

                if latest_shares_outstanding is not None:
                    stock_sheet['C19'].value = latest_shares_outstanding / 1_000_000
                    stock_sheet['C19'].number_format = '0'
                else:
                    stock_sheet['C19'].value = "N/A"
                stock_sheet['C19'].alignment = Alignment(horizontal='center', vertical='center')

                # Calculate Working Capital: latest_current_assets - latest_current_liabilities
                if latest_current_assets is not None and latest_current_liabilities is not None:
                    working_capital = latest_current_assets - latest_current_liabilities
                    stock_sheet['C24'].value = working_capital / 1_000_000
                    stock_sheet['C24'].number_format = '$#,##0'
                else:
                    stock_sheet['C24'].value = "N/A"
                stock_sheet['C24'].alignment = Alignment(horizontal='center', vertical='center')

                # Set formulas for C20, C21, C25, C26
                stock_sheet['C20'].value = "=(C16 - C17 + C18) / C19"
                stock_sheet['C20'].number_format = '$#,##0.00'
                stock_sheet['C20'].alignment = Alignment(horizontal='center', vertical='center')

                stock_sheet['C21'].value = "=C20 / L2"
                stock_sheet['C21'].number_format = '0.0%'
                stock_sheet['C21'].alignment = Alignment(horizontal='center', vertical='center')

                stock_sheet['C25'].value = "=C24 / C19"
                stock_sheet['C25'].number_format = '$#,##0.00'
                stock_sheet['C25'].alignment = Alignment(horizontal='center', vertical='center')

                stock_sheet['C26'].value = "=C25 / L2"
                stock_sheet['C26'].number_format = '0.0%'
                stock_sheet['C26'].alignment = Alignment(horizontal='center', vertical='center')

                # Set conditional formulas and shading for C22 and C27
                stock_sheet['C22'].value = '=IF(C21>=1.5, "Yes", IF(C21>=1, "Maybe", "No"))'
                stock_sheet['C22'].alignment = Alignment(horizontal='center', vertical='center')

                stock_sheet['C27'].value = '=IF(C26>=1.5, "Yes", IF(C26>=1, "Maybe", "No"))'
                stock_sheet['C27'].alignment = Alignment(horizontal='center', vertical='center')

                # Apply conditional formatting for C22 and C27
                green_fill = PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')
                light_blue_fill = PatternFill(start_color='ADD8E6', end_color='ADD8E6', fill_type='solid')
                red_fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')

                stock_sheet.conditional_formatting.add('C22', openpyxl.formatting.rule.CellIsRule(operator='equal', formula=['"Yes"'], fill=green_fill))
                stock_sheet.conditional_formatting.add('C22', openpyxl.formatting.rule.CellIsRule(operator='equal', formula=['"Maybe"'], fill=light_blue_fill))
                stock_sheet.conditional_formatting.add('C22', openpyxl.formatting.rule.CellIsRule(operator='equal', formula=['"No"'], fill=red_fill))

                stock_sheet.conditional_formatting.add('C27', openpyxl.formatting.rule.CellIsRule(operator='equal', formula=['"Yes"'], fill=green_fill))
                stock_sheet.conditional_formatting.add('C27', openpyxl.formatting.rule.CellIsRule(operator='equal', formula=['"Maybe"'], fill=light_blue_fill))
                stock_sheet.conditional_formatting.add('C27', openpyxl.formatting.rule.CellIsRule(operator='equal', formula=['"No"'], fill=red_fill))

                # Skip C23 for spacing
                stock_sheet['C23'].value = ""
                stock_sheet['C23'].alignment = Alignment(horizontal='center', vertical='center')

                # Set thick outside borders for A16:C27
                for row in range(16, 28):
                    for col in ['A', 'B', 'C']:
                        cell = stock_sheet[f'{col}{row}']
                        cell.border = Border(
                            top=bold_side if row == 16 else thin_side,
                            bottom=bold_side if row == 27 else thin_side,
                            left=bold_side if col == 'A' else thin_side,
                            right=bold_side if col == 'C' else thin_side
                        )

                # Set G17, G18, H19, G21, G22, H23, G25, G26, G27
                stock_sheet['G17'].value = "Asset Value Factor Reduction (Page 155)"
                stock_sheet['G17'].font = Font(bold=True)
                stock_sheet['G17'].alignment = Alignment(horizontal='left', vertical='center')

                stock_sheet['G18'].value = "Does Intrinsic Value exceed Value of Cash and Assets by 2?"
                stock_sheet['G18'].alignment = Alignment(horizontal='left', vertical='center')

                stock_sheet['H19'].value = "If Yes... apply Asset Value Reduction to Intrinsic Value."
                stock_sheet['H19'].alignment = Alignment(horizontal='left', vertical='center')

                stock_sheet['G21'].value = "Excess Current Asset Factor (Page 156)"
                stock_sheet['G21'].font = Font(bold=True)
                stock_sheet['G21'].alignment = Alignment(horizontal='left', vertical='center')

                stock_sheet['G22'].value = "Does Value of Cash and Assets exceed Intrinsic Value?"
                stock_sheet['G22'].alignment = Alignment(horizontal='left', vertical='center')

                stock_sheet['H23'].value = "If Yes... apply Current Asset Premium to Intrinsic Value."
                stock_sheet['H23'].alignment = Alignment(horizontal='left', vertical='center')

                stock_sheet['G25'].value = "Price in Relation to Adjusted Intrinsic Value >100%"
                stock_sheet['G25'].font = Font(bold=True)
                stock_sheet['G25'].alignment = Alignment(horizontal='left', vertical='center')

                stock_sheet['G26'].value = "Is the stock a bargain?"
                stock_sheet['G26'].font = Font(bold=True)
                stock_sheet['G26'].alignment = Alignment(horizontal='left', vertical='center')

                stock_sheet['G27'].value = "When should I consider the price to be a bargain?"
                stock_sheet['G27'].font = Font(bold=True)
                stock_sheet['G27'].alignment = Alignment(horizontal='left', vertical='center')

                # Set thin bottom border for G25:L25, G26:L26, G27:L27
                for row in [25, 26, 27]:
                    for col in ['G', 'H', 'I', 'J', 'K', 'L']:
                        cell = stock_sheet[f'{col}{row}']
                        cell.border = Border(bottom=thin_side)

                # Set M18, M19, M22, M23, M25, M26
                stock_sheet['M18'].value = '=IF(M15/C20>=2, "Yes", "No")'
                stock_sheet['M18'].alignment = Alignment(horizontal='center', vertical='center')

                stock_sheet['M19'].value = '=IF(M18="Yes", M15 - (M15 - 2*C20)/4, M15)'
                stock_sheet['M19'].number_format = '$#,##0.00'
                stock_sheet['M19'].alignment = Alignment(horizontal='center', vertical='center')

                stock_sheet['M22'].value = '=IF(C20 - M15 > 0, "Yes", "No")'
                stock_sheet['M22'].alignment = Alignment(horizontal='center', vertical='center')

                stock_sheet['M23'].value = '=IF(M22="Yes", M15 + (C20 - M15)/2, M15)'
                stock_sheet['M23'].number_format = '$#,##0.00'
                stock_sheet['M23'].alignment = Alignment(horizontal='center', vertical='center')

                stock_sheet['M25'].value = '=IF(M18="Yes", M19/L2, IF(M22="Yes", M23/L2, M15/L2))'
                stock_sheet['M25'].number_format = '0.0%'
                stock_sheet['M25'].alignment = Alignment(horizontal='center', vertical='center')

                # Set M26 bargain assessment based on the margin of safety factor 
                stock_sheet['M26'].value = f'=IF(M25>=1+(1-{factor}), "Yes", IF(M25>=1, "Maybe", "No"))'
                stock_sheet['M26'].alignment = Alignment(horizontal='center', vertical='center')

                # Apply conditional formatting for M26
                green_fill = PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')
                light_blue_fill = PatternFill(start_color='ADD8E6', end_color='ADD8E6', fill_type='solid')
                red_fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')

                stock_sheet.conditional_formatting.add('M26', openpyxl.formatting.rule.CellIsRule(operator='equal', formula=['"Yes"'], fill=green_fill))
                stock_sheet.conditional_formatting.add('M26', openpyxl.formatting.rule.CellIsRule(operator='equal', formula=['"Maybe"'], fill=light_blue_fill))
                stock_sheet.conditional_formatting.add('M26', openpyxl.formatting.rule.CellIsRule(operator='equal', formula=['"No"'], fill=red_fill))

                # Set M27 with the updated formula using the margin of safety factor
                stock_sheet['M27'].value = f'=IFERROR(IF(M18="Yes", M19*{factor}, IF(M22="Yes", M23*{factor}, M15*{factor})), "N/A")'
                stock_sheet['M27'].number_format = '$#,##0.00'
                stock_sheet['M27'].alignment = Alignment(horizontal='center', vertical='center')

                # Set A29: "Shares Outstanding"
                stock_sheet['A29'].value = "Shares Outstanding"
                stock_sheet['A29'].font = Font(bold=True)
                stock_sheet['A29'].alignment = Alignment(horizontal='center', vertical='center')

                # Populate B29:K29 with shares outstanding from raw_income_data
                if raw_income_data:
                    income_data_list = json.loads(raw_income_data)
                    shares_outstanding_dict = {}
                    for entry in income_data_list:
                        if 'date' in entry and 'weightedAverageShsOut' in entry:
                            year = entry['date'][:4]
                            shares_outstanding_dict[year] = float(entry['weightedAverageShsOut'])
                    shares_outstanding_list = [shares_outstanding_dict.get(str(year), None) for year in years_list]
                    for col, shares in enumerate(shares_outstanding_list, start=2):  # B to K
                        cell = stock_sheet.cell(row=29, column=col)
                        if shares is not None:
                            cell.value = shares / 1_000_000  # Convert to millions
                            cell.number_format = '0'
                        else:
                            cell.value = "N/A"
                        cell.alignment = Alignment(horizontal='center', vertical='center')

                    # Calculate CAGR for L29 to check if shares are trending downward
                    valid_shares = [s for s in shares_outstanding_list if s is not None and s > 0]
                    if len(valid_shares) >= 2:
                        beginning_value = valid_shares[0]
                        ending_value = valid_shares[-1]
                        n = len(valid_shares) - 1  # Number of periods
                        if beginning_value > 0 and ending_value > 0:
                            cagr = (ending_value / beginning_value) ** (1 / n) - 1
                            stock_sheet['L29'].value = "Yes" if cagr < 0 else "No"
                        else:
                            stock_sheet['L29'].value = "N/A"
                    else:
                        stock_sheet['L29'].value = "N/A"
                    stock_sheet['L29'].alignment = Alignment(horizontal='center', vertical='center')
                else:
                    for col in range(2, 13):  # B to L
                        stock_sheet.cell(row=29, column=col).value = "N/A"
                        stock_sheet.cell(row=29, column=col).alignment = Alignment(horizontal='center', vertical='center')

                # Set A31: "SHARES IF $1000"
                stock_sheet['A31'].value = "SHARES IF $1000"
                stock_sheet['A31'].font = Font(bold=True)
                stock_sheet['A31'].alignment = Alignment(horizontal='center', vertical='center')

                # Set B31: =1000/L2, formatted for numbers with 0 decimals, centered
                stock_sheet['B31'].value = "=1000/L2"
                stock_sheet['B31'].number_format = '0'
                stock_sheet['B31'].alignment = Alignment(horizontal='center', vertical='center')

                # Set A32: "EARNINGS $1000"
                stock_sheet['A32'].value = "EARNINGS $1000"
                stock_sheet['A32'].font = Font(bold=True)
                stock_sheet['A32'].alignment = Alignment(horizontal='center', vertical='center')

                # Set B32 to K32: =$B$31 * B9 for B32, =$B$31 * C9 for C32, etc., formatted as currency, centered
                for col in range(2, 12):  # B to K
                    col_letter = get_column_letter(col)
                    formula = f"=$B$31 * {col_letter}9"
                    cell = stock_sheet.cell(row=32, column=col)
                    cell.value = formula
                    cell.number_format = '$#,##0.00'
                    cell.alignment = Alignment(horizontal='center', vertical='center')

                # Ensure A30 is bold and centered (assuming it's a placeholder or empty)
                stock_sheet['A30'].font = Font(bold=True)
                stock_sheet['A30'].alignment = Alignment(horizontal='center', vertical='center')

                # Center B4:L12 horizontally and vertically
                for row in range(4, 13):
                    for col in range(2, 13):  # B to L
                        cell = stock_sheet.cell(row=row, column=col)
                        cell.alignment = Alignment(horizontal='center', vertical='center')

                # Parse additional data for new sections
                income_data_list = json.loads(raw_income_data) if raw_income_data else []
                net_income_dict = {entry['date'][:4]: float(entry['netIncome']) for entry in income_data_list if 'date' in entry and 'netIncome' in entry}
                net_income_list = [net_income_dict.get(str(year), None) for year in years_list[-10:]] if years_list else []

                balance_data_list = json.loads(balance_data) if balance_data else []
                book_value_dict = {entry['date'][:4]: float(entry['totalStockholdersEquity']) for entry in balance_data_list if 'date' in entry and 'totalStockholdersEquity' in entry}
                book_value_list = [book_value_dict.get(str(year), None) for year in years_list[-10:]] if years_list else []

                key_metrics_list = json.loads(raw_key_metrics_data) if raw_key_metrics_data else []
                pe_ratio_list = [float(entry['peRatio']) for entry in key_metrics_list if 'peRatio' in entry and entry['peRatio'] is not None]
                average_pe = sum(pe_ratio_list) / len(pe_ratio_list) if pe_ratio_list else "N/A"

                # Net Incomes section (A34:K35)
                stock_sheet.merge_cells('A34:A35')
                cell = stock_sheet['A34']
                cell.value = "Net Incomes ($M)"
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

                for col, net_income in enumerate(net_income_list, start=2):  # B34:K34
                    cell = stock_sheet.cell(row=34, column=col)
                    if net_income is not None:
                        cell.value = net_income / 1_000_000
                        cell.number_format = '$#,##0'
                    else:
                        cell.value = "N/A"
                    cell.alignment = Alignment(horizontal='center', vertical='center')

                for col in range(3, 12):  # C35:K35
                    prev_net_income = stock_sheet.cell(row=34, column=col-1).value
                    current_net_income = stock_sheet.cell(row=34, column=col).value
                    if isinstance(prev_net_income, (int, float)) and isinstance(current_net_income, (int, float)) and prev_net_income != 0:
                        change = (current_net_income - prev_net_income) / prev_net_income
                        stock_sheet.cell(row=35, column=col).value = change
                        stock_sheet.cell(row=35, column=col).number_format = '0.00%'
                    else:
                        stock_sheet.cell(row=35, column=col).value = "N/A"
                    stock_sheet.cell(row=35, column=col).alignment = Alignment(horizontal='center', vertical='center')

                # Calculate book_value_per_share_list
                book_value_per_share_list = [
                    bv / so if so and so > 0 else None
                    for bv, so in zip(book_value_list, shares_outstanding_list)
                ]

                # Set A36:A37 merged cell to "Book Value Per Share ($)"
                stock_sheet.merge_cells('A36:A37')
                cell = stock_sheet['A36']
                cell.value = "Book Value Per Share ($)"
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

                # Populate B36:K36 with book_value_per_share_list
                for col, bvps in enumerate(book_value_per_share_list, start=2):  # B to K
                    cell = stock_sheet.cell(row=36, column=col)
                    if bvps is not None:
                        cell.value = bvps
                        cell.number_format = '$#,##0.00'
                    else:
                        cell.value = "N/A"
                    cell.alignment = Alignment(horizontal='center', vertical='center')

                # Calculate % change in C37:K37 for book value per share
                for col in range(3, 12):  # C to K
                    prev_bvps = stock_sheet.cell(row=36, column=col-1).value
                    current_bvps = stock_sheet.cell(row=36, column=col).value
                    if isinstance(prev_bvps, (int, float)) and isinstance(current_bvps, (int, float)) and prev_bvps != 0:
                        change = (current_bvps - prev_bvps) / prev_bvps
                        stock_sheet.cell(row=37, column=col).value = change
                        stock_sheet.cell(row=37, column=col).number_format = '0.00%'
                    else:
                        stock_sheet.cell(row=37, column=col).value = "N/A"
                    stock_sheet.cell(row=37, column=col).alignment = Alignment(horizontal='center', vertical='center')

                # 10 Yr P/E Avg section (A38:B39)
                stock_sheet.merge_cells('A38:A39')
                cell = stock_sheet['A38']
                cell.value = "10 Yr P/E Avg"
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

                stock_sheet.merge_cells('B38:B39')
                cell = stock_sheet['B38']
                if average_pe != "N/A":
                    cell.value = average_pe
                    cell.number_format = '0.00'
                else:
                    cell.value = "N/A"
                cell.alignment = Alignment(horizontal='center', vertical='center')

                # Current Assets section (D38:E39)
                stock_sheet.merge_cells('D38:D39')
                cell = stock_sheet['D38']
                cell.value = "Current Assets ($M)"
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

                stock_sheet.merge_cells('E38:E39')
                cell = stock_sheet['E38']
                if latest_current_assets is not None:
                    cell.value = latest_current_assets / 1_000_000
                    cell.number_format = '$#,##0'
                else:
                    cell.value = "N/A"
                cell.alignment = Alignment(horizontal='center', vertical='center')

                # Short Term Debt section (A40:B41)
                stock_sheet.merge_cells('A40:A41')
                cell = stock_sheet['A40']
                cell.value = "Short Term Debt ($M)"
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

                stock_sheet.merge_cells('B40:B41')
                cell = stock_sheet['B40']
                if latest_short_term_debt is not None:
                    cell.value = latest_short_term_debt / 1_000_000
                    cell.number_format = '$#,##0'
                else:
                    cell.value = "N/A"
                cell.alignment = Alignment(horizontal='center', vertical='center')

                # Current Liabilities section (D40:E41)
                stock_sheet.merge_cells('D40:D41')
                cell = stock_sheet['D40']
                cell.value = "Current Liabilities ($M)"
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

                stock_sheet.merge_cells('E40:E41')
                cell = stock_sheet['E40']
                if latest_current_liabilities is not None:
                    cell.value = latest_current_liabilities / 1_000_000
                    cell.number_format = '$#,##0'
                else:
                    cell.value = "N/A"
                cell.alignment = Alignment(horizontal='center', vertical='center')

                # Long Term Debt section (A42:B43)
                stock_sheet.merge_cells('A42:A43')
                cell = stock_sheet['A42']
                cell.value = "Long Term Debt ($M)"
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

                stock_sheet.merge_cells('B42:B43')
                cell = stock_sheet['B42']
                if latest_long_term_debt is not None:
                    cell.value = latest_long_term_debt / 1_000_000
                    cell.number_format = '$#,##0'
                else:
                    cell.value = "N/A"
                cell.alignment = Alignment(horizontal='center', vertical='center')

                # Book Value (total) section (D42:E43)
                stock_sheet.merge_cells('D42:D43')
                cell = stock_sheet['D42']
                cell.value = "Book Value ($M)"
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

                stock_sheet.merge_cells('E42:E43')
                cell = stock_sheet['E42']
                if latest_book_value is not None:
                    cell.value = latest_book_value / 1_000_000
                    cell.number_format = '$#,##0'
                else:
                    cell.value = "N/A"
                cell.alignment = Alignment(horizontal='center', vertical='center')

                # Benjamin Graham's 10 Rule Stability Test section (A45:A55)
                stability_test_questions = [
                    "Benjamin Graham's 10 Rule Stability Test",
                    "Is the Earnings/Price Ratio >= 2x AAA Bond Yield?",
                    "Is the P/E Ratio <= .4 of highest average P/E of the last 10 years?",
                    "Is the Dividend Yield >= 2/3 AAA Bond Yield?",
                    "Is the Share Price <= 2/3 Book Value Per Share?",
                    "Is the Share Price <= 2/3 Net Current Asset Value (NCAV)?",
                    "Is Total Debt < Book Value?",
                    "Is the Current Ratio >= 2?",
                    "Is Total Debt <= 2*(Current Assets - Current Liabilities - Long Term Debt)?  (NCAV)",
                    "Has Net Income Doubled in the Past 10 years?",
                    "Have Net Incomes declined no more than 5% in the last 10 years?"
                ]

                for row, text in enumerate(stability_test_questions, start=45):
                    cell = stock_sheet[f'A{row}']
                    cell.value = text
                    cell.alignment = Alignment(horizontal='left', vertical='center')
                    if row == 45:
                        cell.font = Font(bold=True)

                # Define center alignment for G45:G55
                alignment = Alignment(horizontal='center', vertical='center')

                # Set formulas for G46 and G47 with error handling
                stock_sheet['G46'].value = '=IF(AND(ISNUMBER(N2), ISNUMBER(L3), N2<>0), IF(1/N2>=2*L3, "Yes", "No"), "N/A")'
                stock_sheet['G47'].value = '=IF(AND(ISNUMBER(N2), ISNUMBER(B38)), IF(N2<=0.4*B38, "Yes", "No"), "N/A")'
                stock_sheet['G48'].value = '=IF(AND(ISNUMBER(K10), ISNUMBER(L2), L2<>0), IF((K10/L2)>=(2/3)*L3, "Yes", "No"), "N/A")'
                stock_sheet['G49'].value = '=IF(AND(ISNUMBER(L2), ISNUMBER(K36)), IF(L2 <= (2/3)*K36, "Yes", "No"), "N/A")'
                stock_sheet['G50'].value = '=IF(AND(ISNUMBER(L2), ISNUMBER(E38), ISNUMBER(E40), ISNUMBER(B42)), IF(L2 <= (2/3)*(E38 - E40 - B42), "Yes", "No"), "N/A")'
                stock_sheet['G51'].value = '=IF(AND(ISNUMBER(B40), ISNUMBER(B42), ISNUMBER(E42)), IF((B40 + B42) < E42, "Yes", "No"), "N/A")'
                stock_sheet['G52'].value = '=IF(AND(ISNUMBER(E38), ISNUMBER(E40), E40<>0), IF(E38/E40 >= 2, "Yes", "No"), "N/A")'
                stock_sheet['G53'].value = '=IF(AND(ISNUMBER(B40), ISNUMBER(B42), ISNUMBER(E38), ISNUMBER(E40)), IF((B40 + B42) <= 2 * (E38 - E40 - B42), "Yes", "No"), "N/A")'
                stock_sheet['G54'].value = '=IF(AND(ISNUMBER(K34), ISNUMBER(B34)), IF(K34 >= 2 * B34, "Yes", "No"), "N/A")'
                stock_sheet['G55'].value = '=IF(COUNT(B34:K34)=10, IF(AND(C34>=0.95*B34, D34>=0.95*C34, E34>=0.95*D34, F34>=0.95*E34, G34>=0.95*F34, H34>=0.95*G34, I34>=0.95*H34, J34>=0.95*I34, K34>=0.95*J34), "Yes", "No"), "N/A")'

                # Set G45 to count the number of "Yes" responses in G46:G55
                stock_sheet['G45'].value = '=COUNTIF(G46:G55, "Yes")'
                stock_sheet['G45'].number_format = '0'

                # Apply center alignment to G45:G55
                alignment = Alignment(horizontal='center', vertical='center')
                for row in range(45, 56):
                    stock_sheet[f'G{row}'].alignment = alignment

                # Add charts if there is data
                if years_list:
                    min_year = min(years_list)
                    max_year = max(years_list)

                    # Add horizontal line data for the first chart (at 12%)
                    stock_sheet['Q2'] = min_year
                    stock_sheet['Q3'] = max_year
                    stock_sheet['R2'] = 0.12
                    stock_sheet['R3'] = 0.12

                    # First Chart: ROE and ROTC
                    chart = ScatterChart()
                    chart.style = 2

                    # ROE series (no data labels)
                    roe_values = Reference(stock_sheet, min_col=2, min_row=5, max_col=11, max_row=5)
                    years_ref = Reference(stock_sheet, min_col=2, min_row=4, max_col=11, max_row=4)
                    roe_series = Series(roe_values, years_ref, title="ROE10")
                    roe_series.marker = Marker('circle')
                    roe_series.graphicalProperties.line.solidFill = "0000FF"  # Blue
                    chart.series.append(roe_series)

                    # ROTC series (no data labels)
                    rotc_values = Reference(stock_sheet, min_col=2, min_row=6, max_col=11, max_row=6)
                    rotc_series = Series(rotc_values, years_ref, title="ROTC10")
                    rotc_series.marker = Marker('circle')
                    rotc_series.graphicalProperties.line.solidFill = "FF0000"  # Red
                    chart.series.append(rotc_series)

                    # Horizontal line at 12% (no data labels)
                    hline_x = Reference(stock_sheet, min_col=17, min_row=2, max_col=17, max_row=3)  # Q2:Q3
                    hline_y = Reference(stock_sheet, min_col=18, min_row=2, max_col=18, max_row=3)  # R2:R3
                    hline_series = Series(hline_y, hline_x, title="12%")
                    hline_series.marker = Marker('none')
                    hline_series.graphicalProperties.line.dashStyle = "dash"
                    hline_series.graphicalProperties.line.solidFill = "000000"  # Black
                    chart.series.append(hline_series)

                    # X-Axis: Years as scale values
                    chart.x_axis.title = "Years"
                    chart.x_axis.number_format = '0'  # Whole numbers for years
                    chart.x_axis.scaling.min = min_year
                    chart.x_axis.scaling.max = max_year
                    chart.x_axis.majorUnit = 1  # One label per year
                    chart.x_axis.tickLblPos = "low"  # Labels at the bottom
                    chart.x_axis.majorGridlines = ChartLines()  # Enable gridlines

                    # Y-Axis: Percentages as scale values
                    chart.y_axis.title = "Percentage"
                    chart.y_axis.number_format = '0%'  # Format as percentage
                    # Calculate appropriate min, max, and majorUnit based on data
                    roe_values_list = [val for val in roe_list if val is not None]
                    rotc_values_list = [val for val in rotc_list if val is not None]
                    all_values = roe_values_list + rotc_values_list + [12]  # Include the 12% line
                    if all_values:
                        min_val = min(all_values) / 100  # Convert percentage to decimal
                        max_val = max(all_values) / 100
                        # Adjust min and max to give some padding
                        min_val = min(min_val, 0)  # Ensure we show at least down to 0%
                        max_val = max_val + 0.05  # Add 5% padding on top
                        # Set a reasonable major unit (e.g., 5% intervals)
                        chart.y_axis.scaling.min = min_val
                        chart.y_axis.scaling.max = max_val
                        range_val = max_val - min_val
                        if range_val > 0:
                            # Aim for around 5-10 tick marks
                            major_unit = max(0.05, round(range_val / 5, 2))  # At least 5% intervals
                            chart.y_axis.majorUnit = major_unit
                    chart.y_axis.majorGridlines = ChartLines()  # Enable gridlines
                    chart.y_axis.tickLblPos = "nextTo"  # Ensure labels are visible

                    # Add chart to sheet
                    stock_sheet.add_chart(chart, "O5")
                    chart.width = 15  # Width in cm
                    chart.height = 8  # Height in cm

                    # Second Chart: EPS and DIV
                    chart2 = ScatterChart()
                    chart2.style = 2

                    # EPS series (no data labels)
                    eps_values = Reference(stock_sheet, min_col=2, min_row=7, max_col=11, max_row=7)
                    eps_series = Series(eps_values, years_ref, title="EPS10")
                    eps_series.marker = Marker('circle')
                    eps_series.graphicalProperties.line.solidFill = "0000FF"  # Blue
                    chart2.series.append(eps_series)

                    # DIV series (no data labels)
                    div_values = Reference(stock_sheet, min_col=2, min_row=10, max_col=11, max_row=10)
                    div_series = Series(div_values, years_ref, title="DIV10")
                    div_series.marker = Marker('circle')
                    div_series.graphicalProperties.line.solidFill = "FF0000"  # Red
                    chart2.series.append(div_series)

                    # X-Axis: Years as scale values
                    chart2.x_axis.title = "Years"
                    chart2.x_axis.number_format = '0'  # Whole numbers for years
                    chart2.x_axis.scaling.min = min_year
                    chart2.x_axis.scaling.max = max_year
                    chart2.x_axis.majorUnit = 1  # One label per year
                    chart2.x_axis.tickLblPos = "low"  # Labels at the bottom
                    chart2.x_axis.majorGridlines = ChartLines()  # Enable gridlines

                    # Y-Axis: Currency as scale values
                    chart2.y_axis.title = "EPS and Dividends"
                    chart2.y_axis.number_format = '$#,##0.00'  # Format as currency
                    # Calculate appropriate min, max, and majorUnit based on data
                    eps_values_list = [val for val in eps_list if val is not None]
                    div_values_list = [val for val in div_list if val is not None]
                    all_values = eps_values_list + div_values_list
                    if all_values:
                        min_val = min(all_values)
                        max_val = max(all_values)
                        # Adjust min and max to give some padding
                        min_val = min(min_val, 0)  # Ensure we show at least down to 0
                        max_val = max_val + (max_val - min_val) * 0.1  # Add 10% padding on top
                        # Set a reasonable major unit (e.g., $1 intervals or based on range)
                        chart2.y_axis.scaling.min = min_val
                        chart2.y_axis.scaling.max = max_val
                        range_val = max_val - min_val
                        if range_val > 0:
                            # Aim for around 5-10 tick marks
                            major_unit = max(1.0, round(range_val / 5, 1))  # At least $1 intervals
                            chart2.y_axis.majorUnit = major_unit
                    chart2.y_axis.majorGridlines = ChartLines()  # Enable gridlines
                    chart2.y_axis.tickLblPos = "nextTo"  # Ensure labels are visible

                    # Add chart to sheet
                    stock_sheet.add_chart(chart2, "O21")
                    chart2.width = 15  # Width in cm
                    chart2.height = 8  # Height in cm

            # Save the workbook
            file_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")], initialfile=f"{exchange}_Qualifying_Stocks.xlsx")
            if file_path:
                wb.save(file_path)
                messagebox.showinfo("Export Successful", f"Qualifying stocks exported to {file_path}")
                analyze_logger.info(f"Exported {exchange} qualifying stocks to {file_path}")
        except Exception as e:
            analyze_logger.error(f"Error exporting {exchange} qualifying stocks: {str(e)}")
            messagebox.showerror("Export Error", f"An error occurred while exporting: {str(e)}")
        finally:
            conn.close()

    def export_nyse_qualifying_stocks(self):
        self.export_qualifying_stocks("NYSE", min_criteria=self.min_criteria_var.get())

    def export_nasdaq_qualifying_stocks(self):
        self.export_qualifying_stocks("NASDAQ", min_criteria=self.min_criteria_var.get())

    def check_file_age(self):
        files = [
            ("NYSE List", NYSE_LIST_FILE),
            ("NASDAQ List", NASDAQ_LIST_FILE),
            ("Favorites", FAVORITES_FILE)
        ]
        messages = []
        for name, file_path in files:
            if os.path.exists(file_path):
                age_days = (time.time() - os.path.getmtime(file_path)) / (24 * 3600)
                messages.append(f"{name}: {age_days:.1f} days old")
            else:
                messages.append(f"{name}: File not found")
        messagebox.showinfo("File Ages", "\n".join(messages))

    def update_ticker_files(self):
        if messagebox.askyesno("Confirm Update", "This will download the latest NYSE and NASDAQ ticker files from FTP. Continue?"):
            def on_complete():
                self.root.after(0, lambda: messagebox.showinfo("Update Complete", "Ticker files have been updated from FTP."))
            def on_error():
                self.root.after(0, lambda: messagebox.showerror("Update Failed", "Failed to download ticker files. Check logs for details."))
            async def update_task():
                try:
                    await self.ticker_manager.download_ticker_files()
                    on_complete()
                except Exception:
                    on_error()
            self.task_queue.put(update_task())
            messagebox.showinfo("Update Started", "Ticker files update has been queued.")

    def clear_cache(self):
        if messagebox.askyesno("Confirm Clear Cache", "This will clear all cached data. Continue?"):
            clear_in_memory_caches()
            conn, cursor = get_stocks_connection()
            try:
                cursor.execute("DELETE FROM stocks")
                cursor.execute("DELETE FROM screening_progress")
                conn.commit()
                messagebox.showinfo("Cache Cleared", "All cached data has been cleared.")
            except sqlite3.Error as e:
                messagebox.showerror("Error", f"Failed to clear database cache: {str(e)}")
            finally:
                conn.close()

    def show_help(self):
        help_text = """
        Graham Screening App Help:

        - **Analyze Stocks**: Enter tickers (e.g., AOS, AAPL) and click "Analyze Stocks" to evaluate them against Graham's criteria.
        - **Favorites**: Save and load ticker lists using the "Save Favorite" and "Favorite Lists" dropdown.
        - **Screening**: Check NYSE or NASDAQ screening boxes and run to find qualifying stocks.
        - **Export**: Export qualifying stocks to Excel after screening.
        - **Progress**: Monitor screening progress and cache usage in the middle panel.
        - **Tabs**: Select a stock in the treeview to see historical data and metrics below.

        For detailed instructions, see the exported Excel's "Start Here" sheet.
        """
        messagebox.showinfo("Help", help_text)

    def filter_tree(self, event) -> None:
        """Filter the treeview based on the search term entered in the search entry widget."""
        search_term = self.search_entry.get().strip().upper()

        # Clear the current treeview contents
        for item in self.tree.get_children():
            self.tree.delete(item)

        # Connect to the database
        conn, cursor = get_stocks_connection()
        try:
            # If there's a search term, filter stocks by ticker or company_name (case-insensitive); otherwise, show all stocks
            if search_term:
                cursor.execute(
                    "SELECT ticker, company_name, common_score FROM stocks WHERE UPPER(ticker) LIKE ? OR UPPER(company_name) LIKE ?",
                    (f"%{search_term}%", f"%{search_term}%")
                )
            else:
                cursor.execute("SELECT ticker, company_name, common_score FROM stocks")

            results = cursor.fetchall()

            # Populate the treeview with filtered results
            for ticker, company_name, common_score in results:
                self.tree.insert(
                    "", "end", text=ticker,
                    values=(company_name, "Unknown", f"{common_score}/6", "", "", "", "")
                )
        except sqlite3.Error as e:
            analyze_logger.error(f"Database error in filter_tree: {str(e)}")
        finally:
            conn.close()

    def safe_insert(self, text_widget, content):
        self.root.after(0, lambda: text_widget.insert(tk.END, content))

    def check_for_updates(self):
        """Perform a startup check for Python and library updates."""
        msg = []
        
        # Check Python version
        latest_python = self.get_latest_python_version()
        current_python = sys.version_info[:3]
        if latest_python and current_python < latest_python:
            current_str = '.'.join(map(str, current_python))
            latest_str = '.'.join(map(str, latest_python))
            msg.append(f"Python version {current_str} is outdated. Latest is {latest_str}. "
                       f"Please download and install from https://www.python.org/downloads/ manually.")
        
        # Check libraries
        outdated_packages = self.get_outdated_packages()
        if outdated_packages is None:
            msg.append("Failed to check for library updates (pip error). Please run 'pip list --outdated' manually.")
        elif outdated_packages:
            msg.append("The following libraries have updates available:\n" + "\n".join(outdated_packages))
        
        if msg:
            full_msg = "\n\n".join(msg)
            if outdated_packages and outdated_packages is not None:  # Offer update only if libraries are outdated
                full_msg += "\n\nWould you like to update the libraries now? (Python update is manual.)"
                if messagebox.askyesno("Updates Available", full_msg):
                    self.update_packages(outdated_packages)
                    messagebox.showinfo("Update Complete", "Libraries updated successfully. Please restart the application.")
                    self.root.quit()  # Quit the app to encourage restart
            else:
                messagebox.showwarning("Updates Recommended", full_msg)

    def get_latest_python_version(self):
        """Fetch the latest Python 3 version from python.org."""
        try:
            response = requests.get("https://www.python.org/downloads/")
            response.raise_for_status()
            # Parse for "Python 3.xx.x" (adjust regex if page structure changes)
            match = re.search(r"Python (\d+\.\d+\.\d+)", response.text)
            if match:
                return tuple(map(int, match.group(1).split('.')))
        except Exception as e:
            analyze_logger.warning(f"Failed to fetch latest Python version: {str(e)}")
        # Fallback to known latest (as of September 2025)
        return (3, 13, 7)

    def get_outdated_packages(self):
        """Get list of outdated pip packages."""
        try:
            shell = (platform.system() == 'Windows')  # Use shell=True on Windows
            # Replace pythonw.exe with python.exe if necessary
            executable = sys.executable
            if executable.endswith('pythonw.exe'):
                executable = executable[:-len('pythonw.exe')] + 'python.exe'
            cmd = [executable, '-m', 'pip', 'list', '--outdated'] if executable else ['pip', 'list', '--outdated']
            result = subprocess.run(cmd, capture_output=True, text=True, check=True, shell=shell)
            outdated = result.stdout.strip()
            if outdated:
                # Parse table output (skip header lines starting with "Package" or "---")
                lines = outdated.splitlines()
                packages = []
                for line in lines[2:]:  # Skip header and separator line
                    if line.strip():  # Ignore empty lines
                        parts = line.split()
                        if parts:  # Ensure there's at least one part (package name)
                            packages.append(parts[0])
                return packages
            return []
        except subprocess.CalledProcessError as e:
            analyze_logger.error(f"Pip subprocess error: Command '{e.cmd}' failed with return code {e.returncode}. Stderr: {e.stderr}")
            return None
        except Exception as e:
            analyze_logger.error(f"Error checking pip updates: {str(e)}")
            return None

    def update_packages(self, packages):
        """Update the specified packages via pip."""
        shell = (platform.system() == 'Windows')  # Use shell=True on Windows
        for pkg in packages:
            try:
                # Replace pythonw.exe with python.exe if necessary
                executable = sys.executable
                if executable.endswith('pythonw.exe'):
                    executable = executable[:-len('pythonw.exe')] + 'python.exe'
                cmd = [executable, '-m', 'pip', 'install', '--upgrade', pkg] if executable else ['pip', 'install', '--upgrade', pkg]
                subprocess.run(cmd, check=True, shell=shell)
                analyze_logger.info(f"Updated package: {pkg}")
            except Exception as e:
                messagebox.showerror("Update Error", f"Failed to update {pkg}: {str(e)}")
                analyze_logger.error(f"Failed to update {pkg}: {str(e)}")

if __name__ == "__main__":
    root = tk.Tk()
    app = GrahamScreeningApp(root)
    root.mainloop()