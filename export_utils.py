import openpyxl
from openpyxl.chart import ScatterChart
from openpyxl.chart.marker import Marker
from openpyxl.chart.axis import ChartLines
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter, quote_sheetname  # Added quote_sheetname
from tkinter import filedialog, messagebox
import json
import pandas as pd
import math
import yfinance as yf
from datetime import datetime
import os
from graham_data import calculate_graham_score_8, get_bank_metrics, calculate_cagr, cache_manager, NYSE_LIST_FILE, NASDAQ_LIST_FILE
from config import graham_logger, FRED_API_KEY
from openpyxl.chart import Reference, Series

bold_side = Side(style='thick')
thin_side = Side(style='thin')
double_side = Side(style='double')

def format_float(value, precision=2):
    if pd.isna(value) or not isinstance(value, (int, float)) or math.isinf(value):
        return "N/A"
    return f"{value:.{precision}f}"

def fetch_current_prices(tickers):
    """Fetch current prices for a list of tickers using yfinance."""
    try:
        data = yf.download(tickers, period="1d", progress=False)['Close']
        if len(tickers) == 1:
            return {tickers[0]: data.iloc[-1] if not data.empty else "N/A"}
        else:
            return {k: v if not pd.isna(v) else "N/A" for k, v in data.iloc[-1].to_dict().items()}
    except Exception as e:
        graham_logger.error(f"Error fetching prices: {str(e)}")
        return {ticker: "N/A" for ticker in tickers}

def calculate_intrinsic_value(stock_dict):
    eps = stock_dict.get('eps_ttm')
    eps_cagr = stock_dict.get('eps_cagr', 0.0)
    if not eps or eps <= 0:
        return float('nan')
    aaa_yield = cache_manager.get_aaa_yield(FRED_API_KEY)
    if aaa_yield <= 0:
        return float('nan')
    g = max(eps_cagr * 100, 0)
    max_multiplier = 15 if stock_dict.get('sector') == "Financials" else 20
    earnings_multiplier = min(8.5 + 2 * g, max_multiplier)
    normalization_factor = 4.4
    value = (eps * earnings_multiplier * normalization_factor) / (100 * aaa_yield)
    return value if not math.isinf(value) and not pd.isna(value) else "N/A"

def get_tangible_book_value_per_share(key_metrics_data: list) -> float:
    """Extract latest tangible book value per share from key_metrics_data."""
    if not key_metrics_data:
        return 0.0
    latest = key_metrics_data[0]
    tbvps = latest.get('tangibleBookValuePerShare', 0.0)
    try:
        return float(tbvps)
    except (ValueError, TypeError):
        return 0.0

def setup_start_here_sheet(start_sheet, margin_of_safety):
    """Setup the 'Start Here' sheet with instructions and constants."""
    start_sheet.sheet_properties.tabColor = "FFFF00"
    for col in range(1, 9):
        start_sheet.column_dimensions[get_column_letter(col)].width = 10
    start_sheet['A1'] = "DIRECTIONS, GENERAL NOTES, CALCULATIONS, AND CONSTANTS FOR STOCK ANALYSIS"
    start_sheet['A1'].font = Font(bold=True, size=16)
    start_sheet['A1'].alignment = Alignment(horizontal='left', vertical='center')
    start_sheet.row_dimensions[1].height = 30

    instructions = [
        "Navigate to the 'Winning Stocks' or 'Financial Winners' sheet and select a company by clicking its name hyperlink to visit its analysis sheet.",
        "Visit Morningstar.com for the stock by clicking the company name hyperlink again. Scroll to 'Financials' and click it. Under 'Financials', click 'Full Financials Data'.",
        "On the 'Full Financials Data' page, select the 'Balance Sheet' tab. Record 'Total Assets' and 'Total Liabilities' for the most recent year.",
        "On the 'Full Financials Data' page, select the 'Key Ratios' tab. From the 'Financials' section, record 'Net Income', 'Earnings Per Share', 'Dividends', 'Shares', and 'Working Capital' for the latest year.",
        "In the 'Key Ratios' tab, under 'Key Ratios' section, record 'Return on Equity' and 'Return on Invested Capital' for the last 10 years.",
        "In the 'Key Ratios' tab, under the 'Financial Health' section, record the most recent 'Long-Term Debt'.",
        "Enter the recorded data into the stock's sheet. The spreadsheet will calculate Graham criteria. Save the file when done."
    ]
    for i, text in enumerate(instructions, start=3):
        start_sheet[f'A{i}'] = text
        start_sheet[f'A{i}'].alignment = Alignment(horizontal='left', vertical='center')

    tips = [
        "High Return on Equity (ROE > 12%) indicates durability; ROE = Net Income / Book Value.",
        "High Return on Total Capital (ROTC > 12%, or >1-1.5% for banks' ROA) shows competitive strength.",
        "Strong, upward-trending Earnings Per Share (EPS = Net Income / Shares Outstanding) signals a durable advantage. Look for sharp drops indicating one-time events.",
        "Long-Term Debt should be less than 5x current net earnings for durability.",
        "Seek companies with products that are repurchased frequently (e.g., Coca-Cola, not airlines).",
        "Avoid companies controlled by labor unions to minimize operational risks.",
        "Favor companies with stable pricing power during economic downturns or inflation.",
        "Choose firms with low reinvestment needs (e.g., H&R Block, not General Motors).",
        "Look for share buybacks, indicating excess cash and confidence in value.",
        "Verify foreign/ADR status (see 'Foreign' tag in app) for currency, regulatory, and tax considerations."
    ]
    for i, tip in enumerate(tips, start=11):
        start_sheet[f'A{i}'] = tip
        start_sheet[f'A{i}'].alignment = Alignment(horizontal='left', vertical='center')

    start_sheet['A22'] = "When to sell? Sell when the stock price exceeds its Intrinsic Value. Compare projected 10-year EPS growth to the return from investing the sale proceeds at the AAA Corporate Bond rate. If the bond yields more, consider selling."
    start_sheet['A22'].alignment = Alignment(horizontal='left', vertical='center')

    start_sheet.merge_cells('A24:B24')
    start_sheet['A24'] = "AAA Corporate Bond Yield"
    start_sheet['A24'].font = Font(bold=True)
    start_sheet['A24'].alignment = Alignment(horizontal='center', vertical='center')
    aaa_yield = cache_manager.get_aaa_yield(FRED_API_KEY)
    start_sheet['C24'] = aaa_yield if aaa_yield is not None else "N/A"
    start_sheet['C24'].number_format = '0.00%'
    start_sheet['C24'].alignment = Alignment(horizontal='center', vertical='center')

    start_sheet.merge_cells('A26:D26')
    start_sheet['A26'] = "Graham Criteria for Investment-Grade Stocks"
    start_sheet['A26'].font = Font(bold=True)
    start_sheet['A26'].alignment = Alignment(horizontal='center', vertical='center')

    investment_criteria = [
        "1) Adequate Size: Annual sales ≥ $500M and assets ≥ $100M.",
        "2) Strong Financial Condition: Current Ratio ≥ 2 (Current Assets / Current Liabilities).",
        "3) Dividend Continuity: At least 10 years of uninterrupted dividends.",
        "4) Earnings Stability: No negative EPS in the last 10 years.",
        "5) Earnings Growth: EPS CAGR ≥ 3% over 10 years.",
        "6) Price-to-Book: Share Price ≤ 1.5x Book Value per Share.",
        "7) Price-to-Earnings: Share Price ≤ 15x average EPS over past 3 years."
    ]
    for i, crit in enumerate(investment_criteria, start=28):
        start_sheet[f'A{i}'] = crit
        start_sheet[f'A{i}'].alignment = Alignment(horizontal='left', vertical='center')

    start_sheet.merge_cells('A36:E36')
    start_sheet['A36'] = "Graham Criteria for Value Stocks (Target: 7/10)"
    start_sheet['A36'].font = Font(bold=True)
    start_sheet['A36'].alignment = Alignment(horizontal='center', vertical='center')

    value_criteria = [
        "1) Earnings Yield ≥ 2x AAA Corporate Bond Yield.",
        "2) P/E Ratio ≤ 40% of highest average P/E over 10 years.",
        "3) Dividend Yield ≥ 2/3 of AAA Bond Yield.",
        "4) Share Price ≤ 2/3 of Tangible Book Value per Share.",
        "5) Share Price ≤ 2/3 of Net Current Asset Value.",
        "6) Total Debt < Tangible Book Value.",
        "7) Current Ratio ≥ 2.",
        "8) Total Debt ≤ 2x Net Current Asset Value.",
        "9) Net Income doubled over the past 10 years.",
        "10) Net Income declines ≤ 5% in any of the last 10 years."
    ]
    for i, crit in enumerate(value_criteria, start=38):
        start_sheet[f'A{i}'] = crit
        start_sheet[f'A{i}'].alignment = Alignment(horizontal='left', vertical='center')

    start_sheet.merge_cells('A48:B48')
    start_sheet['A48'] = "Margin of Safety"
    start_sheet['A48'].font = Font(bold=True)
    start_sheet['A48'].alignment = Alignment(horizontal='center', vertical='center')
    start_sheet['C48'] = margin_of_safety / 100 if margin_of_safety is not None else "N/A"
    start_sheet['C48'].number_format = '0.00%'
    start_sheet['C48'].alignment = Alignment(horizontal='center', vertical='center')

    for row in start_sheet.iter_rows():
        for cell in row:
            if cell.row in [24, 26, 36, 48]:
                cell.alignment = Alignment(wrap_text=False, horizontal='center', vertical='center')
            else:
                cell.alignment = Alignment(wrap_text=False, horizontal='left', vertical='center')

def create_summary_sheet(wb, sheet_name, stocks, is_financial=False, factor=0.0):
    """Create a summary sheet for winning or financial stocks."""
    sheet = wb.create_sheet(sheet_name)
    headers = ["Company Name", "Ticker", "Sector", "Bargain?", "MOS", "Graham Score", "Stability Test", "ROE>12%", "ROTC>12%", "EPS Uptrend", "LTD<5 Years", "Dividend", "Buyback", "POT #1", "POT #2", "Current Price", "Intrinsic Value", "Buy Price"]
    if is_financial:
        headers.extend(["ROA", "ROE", "NIM", "P/TBV"])

    for col, header in enumerate(headers, start=1):
        cell = sheet.cell(row=1, column=col, value=header)
        cell.font = Font(size=12, bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')

    sheet.freeze_panes = 'B2'
    
    last_row = len(stocks) + 1
    for row_idx, stock in enumerate(stocks, start=2):
        ticker_sheet_name = stock['ticker'][:31]  # Unquoted name
        quoted_ticker_sheet_name = quote_sheetname(ticker_sheet_name)  # Quoted if needed
        company_cell = sheet.cell(row=row_idx, column=1, value=stock.get("company_name", "Unknown"))
        company_cell.hyperlink = f"#'{quoted_ticker_sheet_name}'!A1"
        company_cell.style = "Hyperlink"
        
        sheet.cell(row=row_idx, column=2, value=stock.get("ticker", "N/A"))
        sheet.cell(row=row_idx, column=3, value=stock.get("sector", "Unknown"))
        sheet.cell(row=row_idx, column=4, value=f"={quoted_ticker_sheet_name}!M26")
        mos = stock.get("mos", "N/A")
        sheet.cell(row=row_idx, column=5, value=mos / 100 if isinstance(mos, (int, float)) else "N/A")
        sheet.cell(row=row_idx, column=6, value=stock.get("graham_score", "N/A"))
        sheet.cell(row=row_idx, column=7, value=f"={quoted_ticker_sheet_name}!G45")
        sheet.cell(row=row_idx, column=8, value=f"={quoted_ticker_sheet_name}!L5")
        sheet.cell(row=row_idx, column=9, value=f"={quoted_ticker_sheet_name}!L6")
        sheet.cell(row=row_idx, column=10, value=f"=IFERROR(IF(SUMPRODUCT(--({quoted_ticker_sheet_name}!C7:{quoted_ticker_sheet_name}!K7 < {quoted_ticker_sheet_name}!B7:{quoted_ticker_sheet_name}!J7))<=2, \"Yes\", IF(SUMPRODUCT(--({quoted_ticker_sheet_name}!C7:{quoted_ticker_sheet_name}!K7 < {quoted_ticker_sheet_name}!B7:{quoted_ticker_sheet_name}!J7))=3, \"Maybe\", \"No\")), \"N/A\")")
        sheet.cell(row=row_idx, column=11, value=f"={quoted_ticker_sheet_name}!G13")
        sheet.cell(row=row_idx, column=12, value=f"={quoted_ticker_sheet_name}!K10")
        sheet.cell(row=row_idx, column=13, value=f"={quoted_ticker_sheet_name}!L29")
        sheet.cell(row=row_idx, column=14, value=f"={quoted_ticker_sheet_name}!C22")
        sheet.cell(row=row_idx, column=15, value=f"={quoted_ticker_sheet_name}!C27")
        cp = stock.get("current_price", float('nan'))
        if pd.isna(cp):
            sheet.cell(row=row_idx, column=16, value="N/A")
        else:
            sheet.cell(row=row_idx, column=16, value=cp)
            sheet.cell(row=row_idx, column=16).number_format = '$#,##0.00'
        iv = stock.get("intrinsic_value", float('nan'))
        if pd.isna(iv):
            sheet.cell(row=row_idx, column=17, value="N/A")
        else:
            sheet.cell(row=row_idx, column=17, value=iv)
            sheet.cell(row=row_idx, column=17).number_format = '$#,##0.00'
        bp = stock.get("buy_price", float('nan'))
        if pd.isna(bp):
            sheet.cell(row=row_idx, column=18, value="N/A")
        else:
            sheet.cell(row=row_idx, column=18, value=bp)
            sheet.cell(row=row_idx, column=18).number_format = '$#,##0.00'
        
        if is_financial:
            bank_metrics = get_bank_metrics(json.loads(stock.get("raw_key_metrics_data", "[]")))
            tangible_bvps = stock.get("tangible_book_value_per_share", 0)
            price = stock.get("current_price", 1)
            ptbv = price / tangible_bvps if tangible_bvps > 0 else float('nan')
            roa = bank_metrics.get('roa', float('nan'))
            if pd.isna(roa):
                sheet.cell(row=row_idx, column=19, value="N/A")
            else:
                sheet.cell(row=row_idx, column=19, value=roa)
                sheet.cell(row=row_idx, column=19).number_format = '0.00%'
            roe = bank_metrics.get('roe', float('nan'))
            if pd.isna(roe):
                sheet.cell(row=row_idx, column=20, value="N/A")
            else:
                sheet.cell(row=row_idx, column=20, value=roe)
                sheet.cell(row=row_idx, column=20).number_format = '0.00%'
            nim = bank_metrics.get('netInterestMargin', float('nan'))
            if pd.isna(nim):
                sheet.cell(row=row_idx, column=21, value="N/A")
            else:
                sheet.cell(row=row_idx, column=21, value=nim)
                sheet.cell(row=row_idx, column=21).number_format = '0.00%'
            if pd.isna(ptbv):
                sheet.cell(row=row_idx, column=22, value="N/A")
            else:
                sheet.cell(row=row_idx, column=22, value=ptbv)
                sheet.cell(row=row_idx, column=22).number_format = '0.00'

    # Number formats (unchanged)
    for row in range(2, last_row + 1):
        sheet.cell(row=row, column=5).number_format = '0.00%'
        sheet.cell(row=row, column=6).number_format = '0'
        sheet.cell(row=row, column=8).number_format = '0.00%'
        sheet.cell(row=row, column=9).number_format = '0.00%'
        sheet.cell(row=row, column=11).number_format = '0.00'
        sheet.cell(row=row, column=12).number_format = '$#,##0.00'
        # Columns 16-18 already handled above
        if is_financial:
            sheet.cell(row=row, column=19).number_format = '0.00%'
            sheet.cell(row=row, column=20).number_format = '0.00%'
            sheet.cell(row=row, column=21).number_format = '0.00%'
            sheet.cell(row=row, column=22).number_format = '0.00'

    # Conditional formatting (unchanged)
    green_fill = PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')
    light_blue_fill = PatternFill(start_color='ADD8E6', end_color='ADD8E6', fill_type='solid')
    red_fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')

    sheet.conditional_formatting.add(f'D2:D{last_row}', CellIsRule(operator='equal', formula=['"Yes"'], fill=green_fill))
    sheet.conditional_formatting.add(f'D2:D{last_row}', CellIsRule(operator='equal', formula=['"Maybe"'], fill=light_blue_fill))
    sheet.conditional_formatting.add(f'D2:D{last_row}', CellIsRule(operator='equal', formula=['"No"'], fill=red_fill))
    sheet.conditional_formatting.add(f'G2:G{last_row}', FormulaRule(formula=['G2>=8'], fill=green_fill))
    sheet.conditional_formatting.add(f'G2:G{last_row}', FormulaRule(formula=['AND(G2>=6, G2<8)'], fill=light_blue_fill))
    sheet.conditional_formatting.add(f'H2:H{last_row}', FormulaRule(formula=['H2>=0.12'], fill=green_fill))
    sheet.conditional_formatting.add(f'H2:H{last_row}', FormulaRule(formula=['AND(H2>=0.10, H2<0.12)'], fill=light_blue_fill))
    sheet.conditional_formatting.add(f'I2:I{last_row}', FormulaRule(formula=['I2>=0.12'], fill=green_fill))
    sheet.conditional_formatting.add(f'I2:I{last_row}', FormulaRule(formula=['AND(I2>=0.10, I2<0.12)'], fill=light_blue_fill))
    sheet.conditional_formatting.add(f'J2:J{last_row}', CellIsRule(operator='equal', formula=['"Yes"'], fill=green_fill))
    sheet.conditional_formatting.add(f'J2:J{last_row}', CellIsRule(operator='equal', formula=['"Maybe"'], fill=light_blue_fill))
    sheet.conditional_formatting.add(f'J2:J{last_row}', CellIsRule(operator='equal', formula=['"No"'], fill=red_fill))
    sheet.conditional_formatting.add(f'K2:K{last_row}', FormulaRule(formula=['K2<=5'], fill=green_fill))
    sheet.conditional_formatting.add(f'M2:M{last_row}', CellIsRule(operator='equal', formula=['"Yes"'], fill=green_fill))
    sheet.conditional_formatting.add(f'N2:N{last_row}', CellIsRule(operator='equal', formula=['"Yes"'], fill=green_fill))
    sheet.conditional_formatting.add(f'O2:O{last_row}', CellIsRule(operator='equal', formula=['"Yes"'], fill=green_fill))

    column_widths = [55, 12, 25, 14, 10, 19, 18, 14, 16, 18, 17, 14, 14, 12, 12, 18, 20, 15]
    if is_financial:
        column_widths.extend([10, 10, 10, 10])
    for col, width in enumerate(column_widths, start=1):
        sheet.column_dimensions[get_column_letter(col)].width = width

    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center')

    last_col_letter = get_column_letter(len(headers))
    sheet.auto_filter.ref = f"A1:{last_col_letter}{last_row}"
    sheet.auto_filter.add_sort_condition(f"A2:A{last_row}")
    graham_logger.info(f"Created '{sheet_name}' sheet with {len(stocks)} stocks")

def create_stock_sheet(wb, row, prices, factor, exchange):
    """Create an individual stock sheet with data and charts."""
    ticker, company_name, common_score, date, roe, rotc, eps, dividend, ticker_list_hash, balance_data, timestamp, debt_to_equity, eps_ttm, book_value_per_share, latest_revenue, available_data_years, sector, years, latest_total_assets, latest_total_liabilities, latest_shares_outstanding, latest_long_term_debt, latest_short_term_debt, latest_current_assets, latest_current_liabilities, latest_book_value, historic_pe_ratios, latest_net_income, eps_cagr, latest_free_cash_flow, raw_income_data, raw_balance_data, raw_dividend_data, raw_profile_data, raw_cash_flow_data, raw_key_metrics_data = row
    price = prices.get(ticker, "N/A")
    if price == "N/A" or not isinstance(price, (int, float)):
        graham_logger.warning(f"Skipping {ticker} due to invalid price")
        return

    stock_sheet = wb.create_sheet(ticker[:31])

    for row_num in range(1, 56):
        stock_sheet.row_dimensions[row_num].height = 15

    stock_sheet.merge_cells('A2:K3')
    stock_sheet['A2'].value = f"{company_name.upper()} ({ticker})" if company_name else f"Unknown ({ticker})"
    stock_sheet['A2'].font = Font(bold=True, size=18)
    stock_sheet['A2'].alignment = Alignment(horizontal='center', vertical='center')
    stock_sheet['A2'].hyperlink = f"https://www.morningstar.com/stocks/xnys/{ticker}/quote"

    stock_sheet.column_dimensions['A'].width = 30
    for col in range(2, 13):
        stock_sheet.column_dimensions[get_column_letter(col)].width = 10
    stock_sheet.column_dimensions['M'].width = 15
    stock_sheet.column_dimensions['N'].width = 10
    stock_sheet.column_dimensions['O'].width = 15

    ticker_file = NYSE_LIST_FILE if exchange == "NYSE" else NASDAQ_LIST_FILE
    last_updated = datetime.fromtimestamp(os.path.getmtime(ticker_file)).strftime('%d-%b-%y') if os.path.exists(ticker_file) else "Unknown"
    stock_sheet.merge_cells('L1:N1')
    stock_sheet['L1'].value = "Last Updated:"
    stock_sheet['L1'].alignment = Alignment(horizontal='center', vertical='center')
    stock_sheet['O1'].value = last_updated
    stock_sheet['O1'].alignment = Alignment(horizontal='center', vertical='center')

    aaa_yield = cache_manager.get_aaa_yield(FRED_API_KEY)
    pe_ratio = price / eps_ttm if eps_ttm and eps_ttm > 0 else "N/A"

    stock_sheet['L2'].value = price if isinstance(price, (int, float)) else "N/A"
    stock_sheet['L2'].number_format = '$#,##0.00' if isinstance(price, (int, float)) else None
    stock_sheet['L2'].font = Font(bold=True)
    stock_sheet['L2'].alignment = Alignment(horizontal='center', vertical='center')

    stock_sheet['M2'].value = "Current Price"
    stock_sheet['M2'].font = Font(bold=True)
    stock_sheet['M2'].alignment = Alignment(horizontal='center', vertical='center')

    stock_sheet['N2'].value = pe_ratio if isinstance(pe_ratio, (int, float)) else "N/A"
    stock_sheet['N2'].number_format = '0.00' if isinstance(pe_ratio, (int, float)) else None
    stock_sheet['N2'].font = Font(bold=True)
    stock_sheet['N2'].alignment = Alignment(horizontal='center', vertical='center')

    stock_sheet['O2'].value = "P/E Ratio"
    stock_sheet['O2'].font = Font(bold=True)
    stock_sheet['O2'].alignment = Alignment(horizontal='center', vertical='center')

    stock_sheet['L3'].value = aaa_yield if aaa_yield is not None else "N/A"
    stock_sheet['L3'].number_format = '0.00%' if aaa_yield is not None else None
    stock_sheet['L3'].font = Font(bold=True)
    stock_sheet['L3'].alignment = Alignment(horizontal='center', vertical='center')

    stock_sheet.merge_cells('M3:O3')
    stock_sheet['M3'].value = "AAA Corporate Bond Rate"
    stock_sheet['M3'].font = Font(bold=True)
    stock_sheet['M3'].alignment = Alignment(horizontal='center', vertical='center')

    stock_sheet['L2'].border = Border(top=bold_side, left=bold_side, bottom=thin_side)
    stock_sheet['M2'].border = Border(top=bold_side, bottom=thin_side, right=bold_side, left=thin_side)
    stock_sheet['N2'].border = Border(top=bold_side, bottom=thin_side, left=bold_side)
    stock_sheet['O2'].border = Border(top=bold_side, bottom=thin_side, right=bold_side, left=thin_side)
    stock_sheet['L3'].border = Border(top=thin_side, left=bold_side, bottom=bold_side, right=thin_side)
    stock_sheet['M3'].border = Border(top=thin_side, left=thin_side, bottom=bold_side)

    for col in ['N', 'O']:
        stock_sheet[f'{col}4'].border = Border(top=bold_side)

    for col in ['P']:
        stock_sheet[f'{col}3'].border = Border(left=bold_side)

    sub_1 = '\u2081'
    sub_0 = '\u2080'
    sub_10 = sub_1 + sub_0
    labels = [
        "Year",
        f"ROE{sub_10}",
        f"ROTC{sub_10}",
        f"EPS{sub_10}",
        f"EPS{sub_10} CAGR",
        f"EPS{sub_10} Proj",
        f"DIV{sub_10}",
        f"DIV{sub_10} CAGR",
        f"DIV{sub_10} Proj"
    ]
    for i, label in enumerate(labels, start=4):
        stock_sheet[f'A{i}'] = label
        stock_sheet[f'M{i}'] = label
        stock_sheet[f'A{i}'].font = Font(size=10, bold=True)
        stock_sheet[f'M{i}'].font = Font(size=10, bold=True)
        stock_sheet[f'A{i}'].alignment = Alignment(horizontal='center', vertical='center')
        stock_sheet[f'M{i}'].alignment = Alignment(horizontal='center', vertical='center')

    years_list = [int(y) for y in years.split(",") if y.strip().isdigit()] if years else []
    roe_list = [float(x) if x.strip() else None for x in roe.split(",")] if roe else []
    rotc_list = [float(x) if x.strip() else None for x in rotc.split(",")] if rotc else []
    eps_list = [float(x) if x.strip() else None for x in eps.split(",")] if eps else []
    div_list = [float(x) if x.strip() else None for x in dividend.split(",")] if dividend else []

    years_list = years_list[-10:]
    roe_list = roe_list[-10:]
    rotc_list = rotc_list[-10:]
    eps_list = eps_list[-10:]
    div_list = div_list[-10:]

    for col, year in enumerate(years_list, start=2):
        stock_sheet.cell(row=4, column=col).value = year
        stock_sheet.cell(row=4, column=col).font = Font(bold=True)

    for col, roe_value in enumerate(roe_list, start=2):
        cell = stock_sheet.cell(row=5, column=col)
        cell.value = roe_value / 100 if roe_value is not None else "N/A"
        cell.number_format = '0.00%' if roe_value is not None else None

    for col, rotc_value in enumerate(rotc_list, start=2):
        cell = stock_sheet.cell(row=6, column=col)
        cell.value = rotc_value / 100 if rotc_value is not None else "N/A"
        cell.number_format = '0.00%' if rotc_value is not None else None

    for col, eps_value in enumerate(eps_list, start=2):
        cell = stock_sheet.cell(row=7, column=col)
        cell.value = eps_value if eps_value is not None else "N/A"
        cell.number_format = '$#,##0.00' if eps_value is not None else None

    for col, div_value in enumerate(div_list, start=2):
        cell = stock_sheet.cell(row=10, column=col)
        cell.value = div_value if div_value is not None else "N/A"
        cell.number_format = '$#,##0.00' if div_value is not None else None

    stock_sheet['L4'].value = f"Avg{sub_10}"
    stock_sheet['L4'].font = Font(bold=True)
    stock_sheet['L4'].alignment = Alignment(horizontal='center', vertical='center')

    stock_sheet['L5'].value = "=IFERROR(AVERAGE(B5:K5), \"N/A\")"
    stock_sheet['L5'].number_format = '0.00%'
    stock_sheet['L6'].value = "=IFERROR(AVERAGE(B6:K6), \"N/A\")"
    stock_sheet['L6'].number_format = '0.00%'
    stock_sheet['L7'].value = "=IFERROR(AVERAGE(B7:K7), \"N/A\")"
    stock_sheet['L7'].number_format = '$#,##0.00'

    stock_sheet['L8'].value = '=IFERROR(IF(AND(COUNT(B7:K7)>=2, B7<>0), (INDEX(B7:K7, COUNT(B7:K7)) / B7)^(1/(COUNT(B7:K7)-1)) - 1, "N/A"), "N/A")'
    stock_sheet['L8'].number_format = '0.00%'

    stock_sheet['L9'].value = "=IFERROR(AVERAGE(B9:K9), \"N/A\")"
    stock_sheet['L9'].number_format = '$#,##0.00'
    stock_sheet['L10'].value = "=IFERROR(AVERAGE(B10:K10), \"N/A\")"
    stock_sheet['L10'].number_format = '$#,##0.00'

    stock_sheet['L11'].value = '=IFERROR(IF(AND(COUNT(B10:K10)>=2, B10<>0), (INDEX(B10:K10, COUNT(B10:K10)) / B10)^(1/(COUNT(B10:K10)-1)) - 1, "N/A"), "N/A")'
    stock_sheet['L11'].number_format = '0.00%'

    stock_sheet['L12'].value = "=IFERROR(AVERAGE(B12:K12), \"N/A\")"
    stock_sheet['L12'].number_format = '$#,##0.00'

    stock_sheet['B9'].value = "=IFERROR(L7 * (1 + L8), \"N/A\")"
    stock_sheet['B9'].number_format = '$#,##0.00'
    for col in range(3, 12):
        prev_col_letter = get_column_letter(col - 1)
        stock_sheet.cell(row=9, column=col).value = f"=IFERROR({prev_col_letter}9 * (1 + $L$8), \"N/A\")"
        stock_sheet.cell(row=9, column=col).number_format = '$#,##0.00'

    stock_sheet['B12'].value = "=IFERROR(L10 * (1 + L11), \"N/A\")"
    stock_sheet['B12'].number_format = '$#,##0.00'
    for col in range(3, 12):
        prev_col_letter = get_column_letter(col - 1)
        stock_sheet.cell(row=12, column=col).value = f"=IFERROR({prev_col_letter}12 * (1 + $L$11), \"N/A\")"
        stock_sheet.cell(row=12, column=col).number_format = '$#,##0.00'

    stock_sheet['A13'].value = "Long Term Debt ($M)"
    stock_sheet['A13'].font = Font(bold=True)
    stock_sheet['A13'].alignment = Alignment(horizontal='center', vertical='center')
    stock_sheet['B13'].value = latest_long_term_debt / 1_000_000 if latest_long_term_debt is not None else "N/A"
    stock_sheet['B13'].number_format = '$#,##0' if latest_long_term_debt is not None else None

    stock_sheet.merge_cells('C13:D13')
    stock_sheet['C13'].value = "Net Income ($M)"
    stock_sheet['C13'].font = Font(bold=True)
    stock_sheet['C13'].alignment = Alignment(horizontal='center', vertical='center')
    stock_sheet['E13'].value = latest_net_income / 1_000_000 if latest_net_income is not None else "N/A"
    stock_sheet['E13'].number_format = '$#,##0' if latest_net_income is not None else None

    stock_sheet['F13'].value = "LTD/NI"
    stock_sheet['F13'].font = Font(bold=True)
    stock_sheet['F13'].alignment = Alignment(horizontal='center', vertical='center')
    stock_sheet['G13'].value = "=IFERROR(B13/E13, \"N/A\")"
    stock_sheet['G13'].number_format = '0.00'

    for col in ['A', 'B', 'C', 'E', 'F', 'G']:
        stock_sheet[f'{col}13'].alignment = Alignment(horizontal='center', vertical='center')

    for col in ['B', 'C']:
        stock_sheet[f'{col}14'].border = Border(bottom=bold_side)

    for col in ['D']:
        stock_sheet[f'{col}15'].border = Border(left=bold_side)

    stock_sheet['H14'].value = "Intelligent Investor Earnings Multiplier (8.5-20)"
    stock_sheet['H14'].font = Font(bold=True)
    stock_sheet['H14'].alignment = Alignment(horizontal='left', vertical='center')

    g = max(eps_cagr * 100, 0) if eps_cagr is not None else 0
    earnings_multiplier = min(8.5 + 2 * g, 20)
    stock_sheet['M14'].value = earnings_multiplier if earnings_multiplier != "N/A" else "N/A"
    stock_sheet['M14'].number_format = '0.00' if earnings_multiplier != "N/A" else None
    stock_sheet['M14'].alignment = Alignment(horizontal='center', vertical='center')

    for col in ['H', 'I', 'J', 'K', 'L', 'M']:
        stock_sheet[f'{col}14'].border = Border(bottom=thin_side)

    stock_sheet.merge_cells('A15:C15')
    stock_sheet['A15'].value = "Value to a Private Owner Test (POT) (Page 98)"
    stock_sheet['A15'].font = Font(bold=True)
    stock_sheet['A15'].alignment = Alignment(horizontal='center', vertical='center')
    stock_sheet['A15'].border = Border(left=bold_side, top=bold_side, right=bold_side, bottom=bold_side)

    stock_sheet['H15'].value = "Intrinsic Value"
    stock_sheet['H15'].font = Font(bold=True)
    stock_sheet['H15'].alignment = Alignment(horizontal='left', vertical='center')

    stock_sheet['M15'].value = "=IFERROR((K7 * M14 * 4.4) / (100 * L3), \"N/A\")"
    stock_sheet['M15'].number_format = '$#,##0.00'
    stock_sheet['M15'].alignment = Alignment(horizontal='center', vertical='center')

    for col in ['H', 'I', 'J', 'K', 'L', 'M']:
        stock_sheet[f'{col}15'].border = Border(bottom=double_side)

    a_labels = [
        "Total Assets ($M)",
        "Total Liabilities ($M)",
        "Free Cash Flow ($M)",
        "Shares Outstanding (M)",
        "Value Per Share of Cash and Assets:",
        "Price in Relation to POT >100%",
        "Is the Company Worth Owning?",
        "",
        "Working Capital ($M)",
        "Value Per Share of Working Capital:",
        "Price in Relation to Working Capital >100%:",
        "Does Capital Alone Make It A Bargain?"
    ]
    for i, label in enumerate(a_labels, start=16):
        stock_sheet[f'A{i}'] = label
        stock_sheet[f'A{i}'].alignment = Alignment(horizontal='left', vertical='center')

    stock_sheet['C16'].value = latest_total_assets / 1_000_000 if latest_total_assets is not None else "N/A"
    stock_sheet['C16'].number_format = '$#,##0' if latest_total_assets is not None else None
    stock_sheet['C16'].alignment = Alignment(horizontal='center', vertical='center')

    stock_sheet['C17'].value = latest_total_liabilities / 1_000_000 if latest_total_liabilities is not None else "N/A"
    stock_sheet['C17'].number_format = '$#,##0' if latest_total_liabilities is not None else None
    stock_sheet['C17'].alignment = Alignment(horizontal='center', vertical='center')

    stock_sheet['C18'].value = latest_free_cash_flow / 1_000_000 if latest_free_cash_flow is not None else "N/A"
    stock_sheet['C18'].number_format = '$#,##0' if latest_free_cash_flow is not None else None
    stock_sheet['C18'].alignment = Alignment(horizontal='center', vertical='center')

    stock_sheet['C19'].value = latest_shares_outstanding / 1_000_000 if latest_shares_outstanding is not None else "N/A"
    stock_sheet['C19'].number_format = '0' if latest_shares_outstanding is not None else None
    stock_sheet['C19'].alignment = Alignment(horizontal='center', vertical='center')

    working_capital = (latest_current_assets - latest_current_liabilities) / 1_000_000 if latest_current_assets is not None and latest_current_liabilities is not None else "N/A"
    stock_sheet['C24'].value = working_capital
    stock_sheet['C24'].number_format = '$#,##0' if isinstance(working_capital, (int, float)) else None
    stock_sheet['C24'].alignment = Alignment(horizontal='center', vertical='center')

    stock_sheet['C20'].value = "=IFERROR((C16 - C17 + C18) / C19, \"N/A\")"
    stock_sheet['C20'].number_format = '$#,##0.00'
    stock_sheet['C20'].alignment = Alignment(horizontal='center', vertical='center')

    stock_sheet['C21'].value = "=IFERROR(C20 / L2, \"N/A\")"
    stock_sheet['C21'].number_format = '0.0%'
    stock_sheet['C21'].alignment = Alignment(horizontal='center', vertical='center')

    stock_sheet['C25'].value = "=IFERROR(C24 / C19, \"N/A\")"
    stock_sheet['C25'].number_format = '$#,##0.00'
    stock_sheet['C25'].alignment = Alignment(horizontal='center', vertical='center')

    stock_sheet['C26'].value = "=IFERROR(C25 / L2, \"N/A\")"
    stock_sheet['C26'].number_format = '0.0%'
    stock_sheet['C26'].alignment = Alignment(horizontal='center', vertical='center')

    stock_sheet['C22'].value = '=IFERROR(IF(C21>=1.5, "Yes", IF(C21>=1, "Maybe", "No")), "N/A")'
    stock_sheet['C22'].alignment = Alignment(horizontal='center', vertical='center')

    stock_sheet['C27'].value = '=IFERROR(IF(C26>=1.5, "Yes", IF(C26>=1, "Maybe", "No")), "N/A")'
    stock_sheet['C27'].alignment = Alignment(horizontal='center', vertical='center')

    green_fill = PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')
    light_blue_fill = PatternFill(start_color='ADD8E6', end_color='ADD8E6', fill_type='solid')
    red_fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')

    stock_sheet.conditional_formatting.add('C22', CellIsRule(operator='equal', formula=['"Yes"'], fill=green_fill))
    stock_sheet.conditional_formatting.add('C22', CellIsRule(operator='equal', formula=['"Maybe"'], fill=light_blue_fill))
    stock_sheet.conditional_formatting.add('C22', CellIsRule(operator='equal', formula=['"No"'], fill=red_fill))
    stock_sheet.conditional_formatting.add('C27', CellIsRule(operator='equal', formula=['"Yes"'], fill=green_fill))
    stock_sheet.conditional_formatting.add('C27', CellIsRule(operator='equal', formula=['"Maybe"'], fill=light_blue_fill))
    stock_sheet.conditional_formatting.add('C27', CellIsRule(operator='equal', formula=['"No"'], fill=red_fill))

    for row in range(16, 28):
        for col in ['A', 'B', 'C']:
            cell = stock_sheet[f'{col}{row}']
            top = bold_side if row == 16 else thin_side
            bottom = bold_side if row == 27 else thin_side
            left = bold_side if col == 'A' else thin_side
            right = bold_side if col == 'C' else thin_side
            cell.border = Border(top=top, bottom=bottom, left=left, right=right)

    stock_sheet['G17'].value = "Asset Value Factor Reduction (Page 155)"
    stock_sheet['G17'].font = Font(bold=True)
    stock_sheet['G17'].alignment = Alignment(horizontal='left', vertical='center')
    stock_sheet['G18'].value = "Does Intrinsic Value exceed Value of Cash and Assets by 2?"
    stock_sheet['G18'].alignment = Alignment(horizontal='left', vertical='center')
    stock_sheet['H19'].value = "If Yes... apply Asset Value Reduction to Intrinsic Value."
    stock_sheet['H19'].alignment = Alignment(horizontal='left', vertical='center')
    stock_sheet['G21'].value = "Excess Current Asset Factor (Page 156)"
    stock_sheet['G21'].font = Font(bold=True)
    stock_sheet['G21'].alignment = Alignment(horizontal='left', vertical='center')
    stock_sheet['G22'].value = "Does Value of Cash and Assets exceed Intrinsic Value?"
    stock_sheet['G22'].alignment = Alignment(horizontal='left', vertical='center')
    stock_sheet['H23'].value = "If Yes... apply Current Asset Premium to Intrinsic Value."
    stock_sheet['H23'].alignment = Alignment(horizontal='left', vertical='center')
    stock_sheet['G25'].value = "Price in Relation to Adjusted Intrinsic Value >100%"
    stock_sheet['G25'].font = Font(bold=True)
    stock_sheet['G25'].alignment = Alignment(horizontal='left', vertical='center')
    stock_sheet['G26'].value = "Is the stock a bargain?"
    stock_sheet['G26'].font = Font(bold=True)
    stock_sheet['G26'].alignment = Alignment(horizontal='left', vertical='center')
    stock_sheet['G27'].value = "When should I consider the price to be a bargain?"
    stock_sheet['G27'].font = Font(bold=True)
    stock_sheet['G27'].alignment = Alignment(horizontal='left', vertical='center')

    for row in [25, 26, 27]:
        for col in ['G', 'H', 'I', 'J', 'K', 'L']:
            stock_sheet[f'{col}{row}'].border = Border(bottom=thin_side)

    stock_sheet['M18'].value = '=IFERROR(IF(M15/C20>=2, "Yes", "No"), "N/A")'
    stock_sheet['M18'].alignment = Alignment(horizontal='center', vertical='center')

    stock_sheet['M19'].value = '=IFERROR(IF(M18="Yes", M15 - (M15 - 2*C20)/4, M15), "N/A")'
    stock_sheet['M19'].number_format = '$#,##0.00'
    stock_sheet['M19'].alignment = Alignment(horizontal='center', vertical='center')

    stock_sheet['M22'].value = '=IFERROR(IF(C20 - M15 > 0, "Yes", "No"), "N/A")'
    stock_sheet['M22'].alignment = Alignment(horizontal='center', vertical='center')

    stock_sheet['M23'].value = '=IFERROR(IF(M22="Yes", M15 + (C20 - M15)/2, M15), "N/A")'
    stock_sheet['M23'].number_format = '$#,##0.00'
    stock_sheet['M23'].alignment = Alignment(horizontal='center', vertical='center')

    stock_sheet['M25'].value = '=IFERROR(IF(M18="Yes", M19/L2, IF(M22="Yes", M23/L2, M15/L2)), "N/A")'
    stock_sheet['M25'].number_format = '0.0%'
    stock_sheet['M25'].alignment = Alignment(horizontal='center', vertical='center')

    stock_sheet['M26'].value = f'=IFERROR(IF(M25>=1+(1-{factor}), "Yes", IF(M25>=1, "Maybe", "No")), "N/A")'
    stock_sheet['M26'].alignment = Alignment(horizontal='center', vertical='center')

    stock_sheet.conditional_formatting.add('M26', CellIsRule(operator='equal', formula=['"Yes"'], fill=green_fill))
    stock_sheet.conditional_formatting.add('M26', CellIsRule(operator='equal', formula=['"Maybe"'], fill=light_blue_fill))
    stock_sheet.conditional_formatting.add('M26', CellIsRule(operator='equal', formula=['"No"'], fill=red_fill))

    stock_sheet['M27'].value = f'=IFERROR(IF(M18="Yes", M19*{factor}, IF(M22="Yes", M23*{factor}, M15*{factor})), "N/A")'
    stock_sheet['M27'].number_format = '$#,##0.00'
    stock_sheet['M27'].alignment = Alignment(horizontal='center', vertical='center')

    stock_sheet['A29'].value = "Shares Outstanding"
    stock_sheet['A29'].font = Font(bold=True)
    stock_sheet['A29'].alignment = Alignment(horizontal='center', vertical='center')

    if raw_income_data:
        income_data_list = json.loads(raw_income_data)
        shares_outstanding_dict = {entry['date'][:4]: float(entry.get('weightedAverageShsOut', 0)) for entry in income_data_list if 'date' in entry}
        shares_outstanding_list = [shares_outstanding_dict.get(str(year), None) for year in years_list]
        for col, shares in enumerate(shares_outstanding_list, start=2):
            cell = stock_sheet.cell(row=29, column=col)
            cell.value = shares / 1_000_000 if shares is not None else "N/A"
            cell.number_format = '0' if shares is not None else None
            cell.alignment = Alignment(horizontal='center', vertical='center')
       
        valid_shares = [s for s in shares_outstanding_list if s is not None and s > 0]
        if len(valid_shares) >= 2:
            beginning_value = valid_shares[0]
            ending_value = valid_shares[-1]
            n = len(valid_shares) - 1
            if beginning_value > 0 and ending_value > 0:
                cagr = (ending_value / beginning_value) ** (1 / n) - 1
                stock_sheet['L29'].value = "Yes" if cagr < 0 else "No"
            else:
                stock_sheet['L29'].value = "N/A"
        else:
            stock_sheet['L29'].value = "N/A"
        stock_sheet['L29'].alignment = Alignment(horizontal='center', vertical='center')
    else:
        for col in range(2, 13):
            stock_sheet.cell(row=29, column=col).value = "N/A"
            stock_sheet.cell(row=29, column=col).alignment = Alignment(horizontal='center', vertical='center')

    stock_sheet['A31'].value = "SHARES IF $1000"
    stock_sheet['A31'].font = Font(bold=True)
    stock_sheet['A31'].alignment = Alignment(horizontal='center', vertical='center')

    stock_sheet['B31'].value = "=IFERROR(1000/L2, \"N/A\")"
    stock_sheet['B31'].number_format = '0'
    stock_sheet['B31'].alignment = Alignment(horizontal='center', vertical='center')

    stock_sheet['A32'].value = "EARNINGS $1000"
    stock_sheet['A32'].font = Font(bold=True)
    stock_sheet['A32'].alignment = Alignment(horizontal='center', vertical='center')

    for col in range(2, 12):
        col_letter = get_column_letter(col)
        stock_sheet.cell(row=32, column=col).value = f"=IFERROR($B$31 * {col_letter}9, \"N/A\")"
        stock_sheet.cell(row=32, column=col).number_format = '$#,##0.00'
        stock_sheet.cell(row=32, column=col).alignment = Alignment(horizontal='center', vertical='center')

    stock_sheet['A30'].font = Font(bold=True)
    stock_sheet['A30'].alignment = Alignment(horizontal='center', vertical='center')

    for row in range(4, 13):
        for col in range(2, 13):
            stock_sheet.cell(row=row, column=col).alignment = Alignment(horizontal='center', vertical='center')

    income_data_list = json.loads(raw_income_data) if raw_income_data else []
    net_income_dict = {entry['date'][:4]: float(entry.get('netIncome', 0)) for entry in income_data_list if 'date' in entry}
    net_income_list = [net_income_dict.get(str(year), None) for year in years_list[-10:]]

    balance_data_list = json.loads(balance_data) if balance_data else []
    book_value_dict = {entry['date'][:4]: float(entry.get('totalStockholdersEquity', 0)) for entry in balance_data_list if 'date' in entry}
    book_value_list = [book_value_dict.get(str(year), None) for year in years_list[-10:]]

    key_metrics_list = json.loads(raw_key_metrics_data) if raw_key_metrics_data else []
    pe_ratio_list = [float(entry.get('peRatio', 0)) for entry in key_metrics_list if 'peRatio' in entry]
    average_pe = sum(pe_ratio_list) / len(pe_ratio_list) if pe_ratio_list else "N/A"

    stock_sheet.merge_cells('A34:A35')
    stock_sheet['A34'].value = "Net Incomes ($M)"
    stock_sheet['A34'].font = Font(bold=True)
    stock_sheet['A34'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    for col, net_income in enumerate(net_income_list, start=2):
        cell = stock_sheet.cell(row=34, column=col)
        cell.value = net_income / 1_000_000 if net_income is not None else "N/A"
        cell.number_format = '$#,##0' if net_income is not None else None
        cell.alignment = Alignment(horizontal='center', vertical='center')

    for col in range(3, 12):
        prev = stock_sheet.cell(row=34, column=col-1).value
        curr = stock_sheet.cell(row=34, column=col).value
        if isinstance(prev, (int, float)) and isinstance(curr, (int, float)) and prev != 0:
            change = (curr - prev) / prev
            stock_sheet.cell(row=35, column=col).value = change
            stock_sheet.cell(row=35, column=col).number_format = '0.00%'
        else:
            stock_sheet.cell(row=35, column=col).value = "N/A"
        stock_sheet.cell(row=35, column=col).alignment = Alignment(horizontal='center', vertical='center')

    book_value_per_share_list = [bv / so if bv is not None and so is not None and so > 0 else None for bv, so in zip(book_value_list, shares_outstanding_list)]

    stock_sheet.merge_cells('A36:A37')
    stock_sheet['A36'].value = "Book Value Per Share ($)"
    stock_sheet['A36'].font = Font(bold=True)
    stock_sheet['A36'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    for col, bvps in enumerate(book_value_per_share_list, start=2):
        cell = stock_sheet.cell(row=36, column=col)
        cell.value = bvps if bvps is not None else "N/A"
        cell.number_format = '$#,##0.00' if bvps is not None else None
        cell.alignment = Alignment(horizontal='center', vertical='center')

    for col in range(3, 12):
        prev = stock_sheet.cell(row=36, column=col-1).value
        curr = stock_sheet.cell(row=36, column=col).value
        if isinstance(prev, (int, float)) and isinstance(curr, (int, float)) and prev != 0:
            change = (curr - prev) / prev
            stock_sheet.cell(row=37, column=col).value = change
            stock_sheet.cell(row=37, column=col).number_format = '0.00%'
        else:
            stock_sheet.cell(row=37, column=col).value = "N/A"
        stock_sheet.cell(row=37, column=col).alignment = Alignment(horizontal='center', vertical='center')

    stock_sheet.merge_cells('A38:A39')
    stock_sheet['A38'].value = "10 Yr P/E Avg"
    stock_sheet['A38'].font = Font(bold=True)
    stock_sheet['A38'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    stock_sheet.merge_cells('B38:B39')
    stock_sheet['B38'].value = average_pe if average_pe != "N/A" else "N/A"
    stock_sheet['B38'].number_format = '0.00' if average_pe != "N/A" else None
    stock_sheet['B38'].alignment = Alignment(horizontal='center', vertical='center')

    stock_sheet.merge_cells('D38:D39')
    stock_sheet['D38'].value = "Current Assets ($M)"
    stock_sheet['D38'].font = Font(bold=True)
    stock_sheet['D38'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    stock_sheet.merge_cells('E38:E39')
    stock_sheet['E38'].value = latest_current_assets / 1_000_000 if latest_current_assets is not None else "N/A"
    stock_sheet['E38'].number_format = '$#,##0' if latest_current_assets is not None else None
    stock_sheet['E38'].alignment = Alignment(horizontal='center', vertical='center')

    stock_sheet.merge_cells('A40:A41')
    stock_sheet['A40'].value = "Short Term Debt ($M)"
    stock_sheet['A40'].font = Font(bold=True)
    stock_sheet['A40'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    stock_sheet.merge_cells('B40:B41')
    stock_sheet['B40'].value = latest_short_term_debt / 1_000_000 if latest_short_term_debt is not None else "N/A"
    stock_sheet['B40'].number_format = '$#,##0' if latest_short_term_debt is not None else None
    stock_sheet['B40'].alignment = Alignment(horizontal='center', vertical='center')

    stock_sheet.merge_cells('D40:D41')
    stock_sheet['D40'].value = "Current Liabilities ($M)"
    stock_sheet['D40'].font = Font(bold=True)
    stock_sheet['D40'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    stock_sheet.merge_cells('E40:E41')
    stock_sheet['E40'].value = latest_current_liabilities / 1_000_000 if latest_current_liabilities is not None else "N/A"
    stock_sheet['E40'].number_format = '$#,##0' if latest_current_liabilities is not None else None
    stock_sheet['E40'].alignment = Alignment(horizontal='center', vertical='center')

    stock_sheet.merge_cells('A42:A43')
    stock_sheet['A42'].value = "Long Term Debt ($M)"
    stock_sheet['A42'].font = Font(bold=True)
    stock_sheet['A42'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    stock_sheet.merge_cells('B42:B43')
    stock_sheet['B42'].value = latest_long_term_debt / 1_000_000 if latest_long_term_debt is not None else "N/A"
    stock_sheet['B42'].number_format = '$#,##0' if latest_long_term_debt is not None else None
    stock_sheet['B42'].alignment = Alignment(horizontal='center', vertical='center')

    stock_sheet.merge_cells('D42:D43')
    stock_sheet['D42'].value = "Book Value ($M)"
    stock_sheet['D42'].font = Font(bold=True)
    stock_sheet['D42'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    stock_sheet.merge_cells('E42:E43')
    stock_sheet['E42'].value = latest_book_value / 1_000_000 if latest_book_value is not None else "N/A"
    stock_sheet['E42'].number_format = '$#,##0' if latest_book_value is not None else None
    stock_sheet['E42'].alignment = Alignment(horizontal='center', vertical='center')

    stability_test_questions = [
        "Benjamin Graham's 10 Rule Stability Test",
        "Is the Earnings/Price Ratio >= 2x AAA Bond Yield?",
        "Is the P/E Ratio <= .4 of highest average P/E of the last 10 years?",
        "Is the Dividend Yield >= 2/3 AAA Bond Yield?",
        "Is the Share Price <= 2/3 Book Value Per Share?",
        "Is the Share Price <= 2/3 Net Current Asset Value (NCAV)?",
        "Is Total Debt < Book Value?",
        "Is the Current Ratio >= 2?",
        "Is Total Debt <= 2*(Current Assets - Current Liabilities - Long Term Debt)? (NCAV)",
        "Has Net Income Doubled in the Past 10 years?",
        "Have Net Incomes declined no more than 5% in the last 10 years?"
    ]
    for row, text in enumerate(stability_test_questions, start=45):
        stock_sheet[f'A{row}'].value = text
        stock_sheet[f'A{row}'].alignment = Alignment(horizontal='left', vertical='center')
        if row == 45:
            stock_sheet[f'A{row}'].font = Font(bold=True)

    stock_sheet['G46'].value = '=IFERROR(IF(AND(ISNUMBER(N2), ISNUMBER(L3), N2<>0), IF(1/N2>=2*L3, "Yes", "No"), "N/A"), "N/A")'
    stock_sheet['G47'].value = '=IFERROR(IF(AND(ISNUMBER(N2), ISNUMBER(B38)), IF(N2<=0.4*B38, "Yes", "No"), "N/A"), "N/A")'
    stock_sheet['G48'].value = '=IFERROR(IF(AND(ISNUMBER(K10), ISNUMBER(L2), L2<>0), IF((K10/L2)>=(2/3)*L3, "Yes", "No"), "N/A"), "N/A")'
    stock_sheet['G49'].value = '=IFERROR(IF(AND(ISNUMBER(L2), ISNUMBER(K36)), IF(L2 <= (2/3)*K36, "Yes", "No"), "N/A"), "N/A")'
    stock_sheet['G50'].value = '=IFERROR(IF(AND(ISNUMBER(L2), ISNUMBER(E38), ISNUMBER(E40), ISNUMBER(B42)), IF(L2 <= (2/3)*(E38 - E40 - B42), "Yes", "No"), "N/A"), "N/A")'
    stock_sheet['G51'].value = '=IFERROR(IF(AND(ISNUMBER(B40), ISNUMBER(B42), ISNUMBER(E42)), IF((B40 + B42) < E42, "Yes", "No"), "N/A"), "N/A")'
    stock_sheet['G52'].value = '=IFERROR(IF(AND(ISNUMBER(E38), ISNUMBER(E40), E40<>0), IF(E38/E40 >= 2, "Yes", "No"), "N/A"), "N/A")'
    stock_sheet['G53'].value = '=IFERROR(IF(AND(ISNUMBER(B40), ISNUMBER(B42), ISNUMBER(E38), ISNUMBER(E40)), IF((B40 + B42) <= 2 * (E38 - E40 - B42), "Yes", "No"), "N/A"), "N/A")'
    stock_sheet['G54'].value = '=IFERROR(IF(AND(ISNUMBER(K34), ISNUMBER(B34)), IF(K34 >= 2 * B34, "Yes", "No"), "N/A"), "N/A")'
    stock_sheet['G55'].value = '=IFERROR(IF(COUNT(B34:K34)=10, IF(AND(C34>=0.95*B34, D34>=0.95*C34, E34>=0.95*D34, F34>=0.95*E34, G34>=0.95*F34, H34>=0.95*G34, I34>=0.95*H34, J34>=0.95*I34, K34>=0.95*J34), "Yes", "No"), "N/A"), "N/A")'

    stock_sheet['G45'].value = '=COUNTIF(G46:G55, "Yes")'
    stock_sheet['G45'].number_format = '0'

    alignment = Alignment(horizontal='center', vertical='center')
    for row in range(45, 56):
        stock_sheet[f'G{row}'].alignment = alignment

    if years_list:
        min_year = min(years_list) if years_list else 0
        max_year = max(years_list) if years_list else 0
        stock_sheet['Q2'] = min_year
        stock_sheet['Q3'] = max_year
        stock_sheet['R2'] = 0.12
        stock_sheet['R3'] = 0.12

        chart = ScatterChart()
        chart.style = 2

        roe_values = Reference(stock_sheet, min_col=2, min_row=5, max_col=11, max_row=5)
        years_ref = Reference(stock_sheet, min_col=2, min_row=4, max_col=11, max_row=4)
        roe_series = Series(roe_values, years_ref, title="ROE10")
        roe_series.marker = Marker('circle')
        roe_series.graphicalProperties.line.solidFill = "0000FF"
        chart.series.append(roe_series)

        rotc_values = Reference(stock_sheet, min_col=2, min_row=6, max_col=11, max_row=6)
        rotc_series = Series(rotc_values, years_ref, title="ROTC10")
        rotc_series.marker = Marker('circle')
        rotc_series.graphicalProperties.line.solidFill = "FF0000"
        chart.series.append(rotc_series)

        hline_x = Reference(stock_sheet, min_col=17, min_row=2, max_col=17, max_row=3)
        hline_y = Reference(stock_sheet, min_col=18, min_row=2, max_col=18, max_row=3)
        hline_series = Series(hline_y, hline_x, title="12%")
        hline_series.marker = Marker('none')
        hline_series.graphicalProperties.line.dashStyle = "dash"
        hline_series.graphicalProperties.line.solidFill = "000000"
        chart.series.append(hline_series)

        chart.x_axis.title = "Years"
        chart.x_axis.number_format = '0'
        chart.x_axis.scaling.min = min_year
        chart.x_axis.scaling.max = max_year
        chart.x_axis.majorUnit = 1
        chart.x_axis.tickLblPos = "low"
        chart.x_axis.majorGridlines = ChartLines()

        chart.y_axis.title = "Percentage"
        chart.y_axis.number_format = '0%'

        roe_values_list = [val for val in roe_list if val is not None]
        rotc_values_list = [val for val in rotc_list if val is not None]
        all_values = roe_values_list + rotc_values_list + [12]
        if all_values:
            min_val = min(all_values) / 100
            max_val = max(all_values) / 100
            min_val = min(min_val, 0)
            max_val = max_val + 0.05
            chart.y_axis.scaling.min = min_val
            chart.y_axis.scaling.max = max_val
            range_val = max_val - min_val
            if range_val > 0:
                major_unit = max(0.05, round(range_val / 5, 2))
                chart.y_axis.majorUnit = major_unit
        chart.y_axis.majorGridlines = ChartLines()
        chart.y_axis.tickLblPos = "nextTo"

        stock_sheet.add_chart(chart, "O5")
        chart.width = 15
        chart.height = 8

        chart2 = ScatterChart()
        chart2.style = 2

        eps_values = Reference(stock_sheet, min_col=2, min_row=7, max_col=11, max_row=7)
        eps_series = Series(eps_values, years_ref, title="EPS10")
        eps_series.marker = Marker('circle')
        eps_series.graphicalProperties.line.solidFill = "0000FF"
        chart2.series.append(eps_series)

        div_values = Reference(stock_sheet, min_col=2, min_row=10, max_col=11, max_row=10)
        div_series = Series(div_values, years_ref, title="DIV10")
        div_series.marker = Marker('circle')
        div_series.graphicalProperties.line.solidFill = "FF0000"
        chart2.series.append(div_series)

        chart2.x_axis.title = "Years"
        chart2.x_axis.number_format = '0'
        chart2.x_axis.scaling.min = min_year
        chart2.x_axis.scaling.max = max_year
        chart2.x_axis.majorUnit = 1
        chart2.x_axis.tickLblPos = "low"
        chart2.x_axis.majorGridlines = ChartLines()

        chart2.y_axis.title = "EPS and Dividends"
        chart2.y_axis.number_format = '$#,##0.00'

        eps_values_list = [val for val in eps_list if val is not None]
        div_values_list = [val for val in div_list if val is not None]
        all_values = eps_values_list + div_values_list
        if all_values:
            min_val = min(all_values)
            max_val = max(all_values)
            min_val = min(min_val, 0)
            max_val = max_val + (max_val - min_val) * 0.1
            chart2.y_axis.scaling.min = min_val
            chart2.y_axis.scaling.max = max_val
            range_val = max_val - min_val
            if range_val > 0:
                major_unit = max(1.0, round(range_val / 5, 1))
                chart2.y_axis.majorUnit = major_unit
        chart2.y_axis.majorGridlines = ChartLines()
        chart2.y_axis.tickLblPos = "nextTo"

        stock_sheet.add_chart(chart2, "O21")
        chart2.width = 15
        chart2.height = 8

def export_qualifying_stocks(exchange, min_criteria, margin_of_safety, expected_return):
    conn, cursor = cache_manager.get_stocks_connection()
    try:
        cursor.execute("""
            SELECT s.ticker, s.company_name, g.common_score, s.date, s.roe, s.rotc, s.eps, s.dividend,
                s.ticker_list_hash, s.balance_data, s.timestamp, s.debt_to_equity, s.eps_ttm,
                s.book_value_per_share, s.latest_revenue, s.available_data_years, s.sector,
                s.years, s.latest_total_assets, s.latest_total_liabilities, s.latest_shares_outstanding,
                s.latest_long_term_debt, s.latest_short_term_debt, s.latest_current_assets,
                s.latest_current_liabilities, s.latest_book_value, s.historic_pe_ratios,
                s.latest_net_income, s.eps_cagr, s.latest_free_cash_flow, s.raw_income_data,
                s.raw_balance_data, s.raw_dividend_data, s.raw_profile_data, s.raw_cash_flow_data,
                s.raw_key_metrics_data
            FROM stocks s
            JOIN graham_qualifiers g ON s.ticker = g.ticker
            WHERE g.exchange = ? AND g.common_score >= ?
            ORDER BY g.common_score DESC, s.ticker ASC
        """, (exchange, min_criteria))
        results = cursor.fetchall()
        if not results:
            messagebox.showinfo("No Results", f"No {exchange} stocks meet the Graham criteria with minimum score {min_criteria}.")
            graham_logger.info(f"No qualifying {exchange} stocks found with min_criteria={min_criteria}")
            return

        prices = fetch_current_prices([row[0] for row in results])

        wb = openpyxl.Workbook()
        start_sheet = wb.active
        start_sheet.title = "Start Here"
        setup_start_here_sheet(start_sheet, margin_of_safety)

        stock_data_list = []
        for row in results:
            ticker, company_name, common_score, date, roe, rotc, eps, dividend, ticker_list_hash, balance_data, timestamp, debt_to_equity, eps_ttm, book_value_per_share, latest_revenue, available_data_years, sector, years, latest_total_assets, latest_total_liabilities, latest_shares_outstanding, latest_long_term_debt, latest_short_term_debt, latest_current_assets, latest_current_liabilities, latest_book_value, historic_pe_ratios, latest_net_income, eps_cagr, latest_free_cash_flow, raw_income_data, raw_balance_data, raw_dividend_data, raw_profile_data, raw_cash_flow_data, raw_key_metrics_data = row
            price = prices.get(ticker, "N/A")
            if price == "N/A" or not isinstance(price, (int, float)):
                continue

            try:
                eps_list = [float(x) for x in eps.split(",") if x.strip()] if eps else []
                div_list = [float(x) for x in dividend.split(",") if x.strip()] if dividend else []
                balance_data_dict = json.loads(balance_data) if balance_data else []
            except Exception as e:
                graham_logger.error(f"Error parsing data for {ticker}: {str(e)}")
                continue

            tangible_bvps = get_tangible_book_value_per_share(json.loads(raw_key_metrics_data) if raw_key_metrics_data else [])

            intrinsic_value = calculate_intrinsic_value({'eps_ttm': eps_ttm, 'eps_cagr': eps_cagr})
            margin_of_safety_val = (intrinsic_value - price) / intrinsic_value * 100 if not pd.isna(intrinsic_value) and intrinsic_value != 0 else "N/A"

            buy_price = intrinsic_value * (1 - (margin_of_safety / 100)) if not pd.isna(intrinsic_value) else "N/A"

            graham_score = calculate_graham_score_8(ticker, price, None, None, debt_to_equity, eps_list, div_list, {}, balance_data_dict, json.loads(raw_key_metrics_data) if raw_key_metrics_data else [], available_data_years, latest_revenue, sector)

            stock_data = {
                "company_name": company_name or "Unknown",
                "ticker": ticker,
                "sector": sector or "Unknown",
                "mos": margin_of_safety_val,
                "graham_score": graham_score if graham_score is not None else "N/A",
                "current_price": price,
                "intrinsic_value": intrinsic_value,
                "buy_price": buy_price,
                "latest_total_assets": latest_total_assets,
                "latest_total_liabilities": latest_total_liabilities,
                "latest_shares_outstanding": latest_shares_outstanding,
                "latest_current_assets": latest_current_assets,
                "latest_current_liabilities": latest_current_liabilities,
                "free_cash_flow": latest_free_cash_flow,
                "raw_key_metrics_data": raw_key_metrics_data,
                "tangible_book_value_per_share": tangible_bvps
            }
            stock_data_list.append(stock_data)

        stock_data_list.sort(key=lambda x: x["company_name"])

        financial_sectors_lower = ['financial services', 'finance', 'banking', 'financials']
        financial_stocks = [stock for stock in stock_data_list if stock['sector'].lower() in financial_sectors_lower]
        other_stocks = [stock for stock in stock_data_list if stock not in financial_stocks]

        factor_val = 1 - (margin_of_safety / 100)

        if other_stocks:
            create_summary_sheet(wb, "Winning Stocks", other_stocks, is_financial=False, factor=factor_val)

        if financial_stocks:
            create_summary_sheet(wb, "Financial Winners", financial_stocks, is_financial=True, factor=factor_val)

        results.sort(key=lambda x: x[0])
        for row in results:
            create_stock_sheet(wb, row, prices, factor_val, exchange)

        file_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")], initialfile=f"{exchange}_Qualifying_Stocks.xlsx")
        if file_path:
            wb.save(file_path)
            messagebox.showinfo("Export Successful", f"Qualifying stocks exported to {file_path}")
            graham_logger.info(f"Exported {exchange} qualifying stocks to {file_path}")
    except Exception as e:
        graham_logger.error(f"Error exporting {exchange} qualifying stocks: {str(e)}")
        messagebox.showerror("Export Error", f"An error occurred while exporting: {str(e)}")
    finally:
        conn.close()